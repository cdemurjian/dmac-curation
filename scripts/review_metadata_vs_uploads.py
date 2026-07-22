#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
# Lifted from lee/scripts/review_nar_vs_uploads.py; renamed and generalized
# for dmac-curation plugin. Lee-specific ACTIVE_UPLOADS map replaced with
# auto-discovery from assay_sheets/. Hardcoded metadata path replaced with
# --metadata-xlsx arg.
"""Comprehensive review: metadata workbook vs assay_sheets/*-upload*.xlsx.

  1) For each upload sheet found in assay_sheets/:
       - row-count delta vs the corresponding sheet in the metadata workbook
       - UIDs only in metadata / only in upload sheet / common
       - For common UIDs: per-field diff on key columns
         (File_PrimaryData, Link_PrimaryData, Parent, Accession, Checksum_PrimaryData)
       - Link_PrimaryData fill rate + URL-type tally (zenodo/omero/geo/other)

  2) Unique Protocol values across all metadata sheets — printed at end, deduplicated.

Usage:
  uv run scripts/review_metadata_vs_uploads.py --metadata-xlsx path/to/All-Metadata.xlsx
  uv run scripts/review_metadata_vs_uploads.py   # auto-discovers previous_metadata/*All*.xlsx

Output is a single readable report on stdout.
"""
from __future__ import annotations
import argparse
import re
import sys
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _config import add_config_args, config_from_args  # noqa: E402
from _config import ProjectRootError  # noqa: E402

COMPARE_COLS = [
    "File_PrimaryData", "Link_PrimaryData", "Parent",
    "Accession", "Checksum_PrimaryData",
]

# Sample types chat_nextseek auto-pulls when resolving lineage. Their presence
# in a download is expected, not "extra rows" worth alarming about.
AUTO_PULLED_PARENT_TYPES = {"MUS", "TIS", "DNA", "RNA", "PAT", "PAV", "CHM", "CEL"}


def load_retrieve_uids(path):
    """Read RETRIEVE.TXT into a set of UIDs. None when the file is absent.

    PHASES.md named this as a Phase 12 input while the script had no flag to
    read it, so the documented diff never ran.
    """
    path = Path(path)
    if not path.is_file():
        return None
    return {line.strip() for line in path.read_text().splitlines() if line.strip()}


def diff_retrieve(requested, downloaded, parent_types=AUTO_PULLED_PARENT_TYPES):
    """Classify the round trip.

    Args:
      requested:    UIDs from RETRIEVE.TXT, or None to skip entirely.
      downloaded:   UIDs present in the downloaded workbook.
      parent_types: sample-type prefixes auto-pulled by lineage.

    Returns:
      None when `requested` is None, else
      {"missing": [...],              # asked for, not in the download
       "auto_pulled_parents": [...],  # in the download, expected, not asked for
       "extra": [...]}                # in the download, unexpected
    """
    if requested is None:
        return None
    missing = sorted(requested - downloaded)
    unrequested = downloaded - requested
    auto = sorted(u for u in unrequested if u.split("-", 1)[0] in parent_types)
    extra = sorted(u for u in unrequested if u.split("-", 1)[0] not in parent_types)
    return {"missing": missing, "auto_pulled_parents": auto, "extra": extra}


# TODO(v0.2): support a project-level ACTIVE_UPLOADS JSON override to handle
# multi-sheet merges (e.g. D.IMG has 4 upload sheets). Auto-discovery uses
# a simple stype -> single sheet mapping.


def discover_active_uploads(sheets_dir: Path) -> dict[str, list[str]]:
    """Auto-discover sample-type → [upload sheet filenames] from assay_sheets/ glob."""
    result: dict[str, list[str]] = {}
    for p in sorted(sheets_dir.glob("*-upload*.xlsx")):
        stem = p.stem  # e.g. "A.FLOW-upload"
        stype = stem.split("-upload")[0]
        if stype:
            result.setdefault(stype, []).append(p.name)
    return result


def load_sheet_rows(path: Path, sheet_name: str | None = None) -> tuple[list[str], list[dict]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else (wb["Samples"] if "Samples" in wb.sheetnames else wb.active)
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], []
        hdr = [str(c) if c is not None else "" for c in rows[0]]
        data = []
        for r in rows[1:]:
            if not r or not r[0]:
                continue
            d = {k: v for k, v in zip(hdr, r)}
            data.append(d)
        return hdr, data
    finally:
        wb.close()


def url_type(s: str | None) -> str:
    if not s:
        return "empty"
    s = str(s).lower()
    if "zenodo.org" in s: return "zenodo"
    if "omero" in s: return "omero"
    if "geo/query" in s or "ncbi.nlm.nih.gov/geo" in s: return "geo"
    if s.startswith("http"): return "other_url"
    return "non_url"


def normalize(v):
    if v is None: return ""
    return str(v).strip()


def compare_sheet(meta: Path, sheets_dir: Path, stype: str, upload_files: list[str]) -> None:
    print(f"\n{'='*78}")
    print(f"  {stype}")
    print(f"{'='*78}")

    # Metadata side
    meta_hdr, meta_rows = load_sheet_rows(meta, stype)
    meta_uids = {r.get("UID") for r in meta_rows if r.get("UID")}
    meta_map = {r["UID"]: r for r in meta_rows if r.get("UID")}

    # Upload side — merge multiple sheets if needed (later overrides earlier on UID conflict)
    up_map: dict[str, dict] = {}
    up_hdrs_seen: set[str] = set()
    for fname in upload_files:
        path = sheets_dir / fname
        if not path.exists():
            print(f"  WARNING: upload sheet not found: {fname}")
            continue
        h, rows = load_sheet_rows(path)
        for col in h:
            up_hdrs_seen.add(col)
        for r in rows:
            if r.get("UID"):
                up_map[r["UID"]] = r
    up_uids = set(up_map.keys())

    print(f"  metadata rows:    {len(meta_rows)}")
    print(f"  upload rows:      {len(up_map)}  (from {', '.join(upload_files)})")

    only_meta = meta_uids - up_uids
    only_up = up_uids - meta_uids
    common = meta_uids & up_uids
    print(f"  common UIDs:      {len(common)}")
    print(f"  only in metadata: {len(only_meta)}")
    print(f"  only in upload:   {len(only_up)}")

    if only_meta and len(only_meta) <= 10:
        print(f"    metadata-only: {sorted(only_meta)}")
    elif only_meta:
        print(f"    metadata-only (first 10 of {len(only_meta)}): {sorted(only_meta)[:10]}")

    if only_up and len(only_up) <= 10:
        print(f"    upload-only: {sorted(only_up)}")
    elif only_up:
        print(f"    upload-only (first 10 of {len(only_up)}): {sorted(only_up)[:10]}")

    # Per-field diff for common UIDs
    diffs_by_col: dict[str, list[tuple[str, str, str]]] = {c: [] for c in COMPARE_COLS}
    for uid in common:
        n = meta_map[uid]
        u = up_map[uid]
        for col in COMPARE_COLS:
            if col not in n and col not in u:
                continue
            nv = normalize(n.get(col))
            uv = normalize(u.get(col))
            if nv != uv:
                diffs_by_col[col].append((uid, nv, uv))

    for col, lst in diffs_by_col.items():
        if not lst: continue
        print(f"\n  diff {col}: {len(lst)} differences")
        for uid, nv, uv in lst[:5]:
            nv_show = (nv[:60] + "...") if len(nv) > 60 else nv
            uv_show = (uv[:60] + "...") if len(uv) > 60 else uv
            print(f"     {uid}")
            print(f"       metadata: {nv_show!r}")
            print(f"       upload:   {uv_show!r}")
        if len(lst) > 5:
            print(f"     ... +{len(lst) - 5} more")

    # Link fill rate / URL type tally (metadata side, source of truth for FDH push)
    if "Link_PrimaryData" in meta_hdr:
        type_counts = Counter(url_type(r.get("Link_PrimaryData")) for r in meta_rows)
        types_str = ", ".join(f"{t}={c}" for t, c in sorted(type_counts.items()))
        filled = sum(c for t, c in type_counts.items() if t != "empty")
        print(f"\n  metadata Link_PrimaryData: {filled}/{len(meta_rows)} filled  ({types_str})")
    elif "Link_PrimaryData" in up_hdrs_seen:
        print(f"\n  metadata sheet has no Link_PrimaryData column (schema needs update on FDH)")


def collect_protocols(meta: Path) -> dict[str, int]:
    """Return {protocol: count} aggregated across all metadata sheets."""
    counts: Counter[str] = Counter()
    sources: dict[str, set[str]] = {}
    wb = load_workbook(meta, read_only=True, data_only=True)
    try:
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            if not rows: continue
            hdr = [str(c) if c is not None else "" for c in rows[0]]
            if "Protocol" not in hdr:
                continue
            i = hdr.index("Protocol")
            for r in rows[1:]:
                if not r or len(r) <= i:
                    continue
                p = r[i]
                if p is None or str(p).strip() == "":
                    continue
                # Split semicolon-separated protocol lists
                for tok in re.split(r"[;]+", str(p)):
                    tok = tok.strip()
                    if not tok: continue
                    counts[tok] += 1
                    sources.setdefault(tok, set()).add(sname)
    finally:
        wb.close()
    # Print
    print(f"\n{'='*78}")
    print(f"  UNIQUE PROTOCOLS  ({len(counts)} distinct)")
    print(f"{'='*78}")
    for proto, c in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])):
        src = ", ".join(sorted(sources[proto]))
        print(f"  [{c:4}x]  {proto}")
        print(f"           seen in: {src}")
    return dict(counts)


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    add_config_args(ap)
    ap.add_argument(
        "--metadata-xlsx",
        metavar="XLSX",
        help="Path to All-Metadata workbook (default: newest previous_metadata/*All*.xlsx)",
    )
    ap.add_argument(
        "--assay-sheets",
        type=Path,
        default=None,
        metavar="DIR",
        help="Directory of upload sheets (default: <project-root>/assay_sheets)",
    )
    ap.add_argument(
        "--retrieve", type=Path, default=None,
        help="RETRIEVE.TXT of requested UIDs. Reports which were requested but "
             "absent from the download (default: <project-root>/RETRIEVE.TXT "
             "when present)",
    )
    args = ap.parse_args()

    try:
        cfg = config_from_args(args)
    except ProjectRootError as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(2)

    meta = Path(args.metadata_xlsx).resolve() if args.metadata_xlsx else cfg.master_workbook
    if meta is None or not meta.exists():
        print("ERROR: no metadata xlsx found. Pass --metadata-xlsx <path>.", file=sys.stderr)
        sys.exit(1)
    sheets_dir = Path(args.assay_sheets).resolve() if args.assay_sheets else cfg.assay_sheets
    active_uploads = discover_active_uploads(sheets_dir)

    print(f"REVIEW: metadata vs assay_sheets/")
    print(f"  metadata: {meta.name}")
    print(f"  sheets dir: {sheets_dir}")
    print(f"  discovered {len(active_uploads)} sample types\n")

    for stype, files in sorted(active_uploads.items()):
        compare_sheet(meta, sheets_dir, stype, files)

    # Check for metadata sheets with no corresponding upload sheet, and collect
    # every UID present in the downloaded workbook (all sheets) for the RETRIEVE
    # round-trip diff below.
    downloaded_uids: set[str] = set()
    wb = load_workbook(meta, read_only=True, data_only=True)
    try:
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                hdr = [str(c) if c is not None else "" for c in rows[0]]
                uid_i = hdr.index("UID") if "UID" in hdr else None
                if uid_i is not None:
                    for r in rows[1:]:
                        if not r or not r[0] or uid_i >= len(r):
                            continue
                        u = r[uid_i]
                        if u is not None and str(u).strip():
                            downloaded_uids.add(str(u).strip())
            if sname not in active_uploads:
                n = sum(1 for r in rows[1:] if r and r[0])
                if n > 0:
                    print(f"\n{'='*78}")
                    print(f"  {sname}  (no local upload sheet)")
                    print(f"{'='*78}")
                    print(f"  metadata rows: {n}")
    finally:
        wb.close()

    # RETRIEVE round trip: PHASES.md named RETRIEVE.TXT as a Phase 12 input, but
    # the script never read it. --retrieve defaults to the project's RETRIEVE.TXT
    # when present, and is skipped with a printed note when absent.
    retrieve_path = args.retrieve or (cfg.root / "RETRIEVE.TXT")
    requested = load_retrieve_uids(retrieve_path)
    if requested is None:
        print(f"\nRETRIEVE round trip: skipped (no {retrieve_path})")
    else:
        d = diff_retrieve(requested, downloaded_uids)
        print("\n" + "-" * 60)
        print("RETRIEVE ROUND TRIP")
        print("-" * 60)
        print(f"Source: {retrieve_path}")
        print(f"Requested: {len(requested)}   Downloaded: {len(downloaded_uids)}")
        print(f"  auto-pulled parents (expected): {len(d['auto_pulled_parents'])}")
        if d["missing"]:
            print(f"  REQUESTED BUT MISSING: {len(d['missing'])}")
            for u in d["missing"][:20]:
                print(f"      - {u}")
            if len(d["missing"]) > 20:
                print(f"      ... and {len(d['missing']) - 20} more")
        else:
            print("  every requested UID is present")
        if d["extra"]:
            print(f"  unexpected extra rows: {len(d['extra'])}")
            for u in d["extra"][:20]:
                print(f"      - {u}")

    collect_protocols(meta)


if __name__ == "__main__":
    main()
