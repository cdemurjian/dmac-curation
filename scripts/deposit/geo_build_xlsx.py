#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
# Lifted from lee/scripts/render_geo_xlsx.py; renamed geo_build_xlsx.py for dmac-curation plugin.
"""Render a filled GEO JSON into an existing GEO template xlsx.

Usage:
  uv run scripts/deposit/geo_build_xlsx.py BULK_filled.json BULK.xlsx BULK_filled.xlsx
  uv run scripts/deposit/geo_build_xlsx.py SPTX_filled.json SPTX.xlsx SPTX_filled.xlsx

Behavior:
  - Loads template (preserves Instructions, EXAMPLEs, Data validation, MD5 Checksums sheets).
  - Loads JSON.
  - In Metadata sheet:
      - Captures the static block from PROTOCOLS row through the PE file-name-1 header
        row so it can be re-pasted below the (possibly expanded) SAMPLES block.
      - Wipes from samples_hdr_row to end of sheet.
      - Writes SAMPLES header + sample rows starting at samples_hdr_row.
      - Leaves a 1-row gap, then re-pastes the PROTOCOLS+PE-header block.
      - Writes PAIRED-END EXPERIMENTS data rows after the re-pasted file-name-1 header.
  - STUDY rows and the PROTOCOLS/PE block itself are preserved verbatim (just shifted).
"""
from __future__ import annotations
import argparse
import json
import sys
from pathlib import Path
from openpyxl import load_workbook

# How many columns to capture/re-paste for the static PROTOCOLS+PE block
STATIC_BLOCK_COLS = 8


def find_row_by_first_col(ws, value):
    """Return 1-based row index where col A == value, else None."""
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[0] == value:
            return i
    return None


def render(json_path: Path, template_path: Path, out_path: Path):
    print(f"Loading template: {template_path}")
    wb = load_workbook(template_path)
    ws = wb["Metadata"]

    with open(json_path) as f:
        data = json.load(f)
    samples = data["samples"]
    paired = data.get("paired_end_experiments", [])

    # Find header rows in template
    samples_hdr_row = find_row_by_first_col(ws, "*library name")
    protocols_row = find_row_by_first_col(ws, "PROTOCOLS")
    pe_hdr_row = find_row_by_first_col(ws, "file name 1")

    if samples_hdr_row is None:
        raise ValueError("Couldn't find SAMPLES header row (*library name) in template")
    if protocols_row is None:
        raise ValueError("Couldn't find PROTOCOLS row in template")
    if pe_hdr_row is None:
        raise ValueError("Couldn't find PAIRED-END EXPERIMENTS header row (file name 1) in template")
    print(
        f"SAMPLES header row {samples_hdr_row}, "
        f"PROTOCOLS row {protocols_row}, PE header row {pe_hdr_row}"
    )

    # --- Capture static block: PROTOCOLS row through PE file-name header row, inclusive ---
    static_block = []
    for r in range(protocols_row, pe_hdr_row + 1):
        static_block.append(
            [ws.cell(row=r, column=c).value for c in range(1, STATIC_BLOCK_COLS + 1)]
        )
    print(f"Captured {len(static_block)} rows of PROTOCOLS+PE-header block")

    # --- Determine output columns for SAMPLES ---
    base_cols = [
        "*library name", "*title", "*library strategy", "*organism",
        "**tissue", "**cell line", "**cell type",
        "genotype", "treatment", "batch",
        "*molecule", "*single or paired-end", "*instrument model", "description",
    ]

    def raw_count(s):
        if "raw_files" in s:
            return len(s["raw_files"])
        return 1

    def proc_count(s):
        if "processed_data_files" in s:
            return len(s["processed_data_files"])
        return sum(
            1 for k in ("processed data file", "processed data file (2)") if s.get(k)
        )

    max_raw = max((raw_count(s) for s in samples), default=1)
    max_proc = max((proc_count(s) for s in samples), default=0)

    proc_headers = []
    if max_proc >= 1:
        proc_headers.append("processed data file")
    for i in range(2, max_proc + 1):
        proc_headers.append(f"processed data file ({i})")

    raw_headers = ["*raw file"]
    if max_raw >= 2:
        raw_headers.append("raw file")
    for i in range(2, max_raw):
        raw_headers.append(f"raw file ({i})")

    full_header = base_cols + proc_headers + raw_headers
    print(f"Writing {len(full_header)} columns: {full_header}")

    # --- Wipe from samples_hdr_row to end of sheet ---
    old_max_row = ws.max_row
    old_max_col = ws.max_column
    for r in range(samples_hdr_row, old_max_row + 1):
        for c in range(1, old_max_col + 1):
            ws.cell(row=r, column=c).value = None

    # --- Write SAMPLES header ---
    for i, h in enumerate(full_header, start=1):
        ws.cell(row=samples_hdr_row, column=i).value = h

    # --- Write sample rows ---
    sample_start_row = samples_hdr_row + 1
    for ri, s in enumerate(samples, start=sample_start_row):
        row_vals = [s.get(col) for col in base_cols]

        # Processed
        proc_list = s.get("processed_data_files") or []
        if not proc_list:
            if s.get("processed data file"):
                proc_list = [s["processed data file"]]
            if s.get("processed data file (2)"):
                proc_list.append(s["processed data file (2)"])
        for i in range(len(proc_headers)):
            row_vals.append(proc_list[i] if i < len(proc_list) else None)

        # Raw
        raw_list = s.get("raw_files") or [v for v in [s.get("*raw file")] if v]
        for i in range(len(raw_headers)):
            row_vals.append(raw_list[i] if i < len(raw_list) else None)

        for ci, v in enumerate(row_vals, start=1):
            ws.cell(row=ri, column=ci).value = v

    last_sample_row = sample_start_row + len(samples) - 1

    # --- Re-paste static PROTOCOLS+PE-header block ---
    # Leave 1 blank row gap (matches template style: blank row 53 before PROTOCOLS).
    static_start_row = last_sample_row + 2
    for offset, row_vals in enumerate(static_block):
        for ci, v in enumerate(row_vals, start=1):
            if v is not None:
                ws.cell(row=static_start_row + offset, column=ci).value = v
    new_pe_hdr_row = static_start_row + (len(static_block) - 1)

    # --- Write PAIRED-END EXPERIMENTS data rows ---
    for ri, pe in enumerate(paired, start=new_pe_hdr_row + 1):
        ws.cell(row=ri, column=1).value = pe.get("file name 1")
        ws.cell(row=ri, column=2).value = pe.get("file name 2")
        ws.cell(row=ri, column=3).value = pe.get("file name 3")
        ws.cell(row=ri, column=4).value = pe.get("file name 4")

    print(
        f"Wrote {len(samples)} sample rows (rows {sample_start_row}-{last_sample_row}) "
        f"and {len(paired)} paired-end rows (after row {new_pe_hdr_row})"
    )
    print(
        f"PROTOCOLS block now at row {static_start_row}, "
        f"PE header now at row {new_pe_hdr_row}"
    )
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("json_path", type=Path, metavar="JSON", help="Filled GEO JSON file")
    ap.add_argument("template_path", type=Path, metavar="TEMPLATE", help="GEO template xlsx")
    ap.add_argument("out_path", type=Path, metavar="OUTPUT", help="Output xlsx path")
    args = ap.parse_args()
    render(args.json_path, args.template_path, args.out_path)


if __name__ == "__main__":
    main()
