#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""
QA pass over the consolidated flat-format xlsx for IntravChip.

Default target:
  assay_sheets/IntravChip_upload.xlsx   (single "Samples" sheet)

Checks (Charlie-level, would-Yufei-notice):
  1. UID uniqueness within the consolidated file
  2. Sampletype validity (must be in context/sampletypes_db.json)
  3. Parent UID resolvability:
       - matches another UID in the upload sheet, OR
       - matches a UID in previous_metadata/MetNet All 260527.xlsx
         (the existing-MetNet baseline — IntravChip OOC rows have CEL
         parents that were registered in prior MetNet batches), OR
       - is intentionally blank for root sampletypes (CEL stock, MDL CAD)
  4. json_metadata cell parses as valid JSON
  5. Required-metadata coverage per sampletype (from sampletypes_db.json)
       — fields blank are flagged but most aren't all-or-nothing blockers
  6. Name uniqueness within sampletype (best-effort; some names may be NULL
     because IntravChip rows lean on UID for identity)
  7. Placeholder sniff: count '*** PLACEHOLDER' markers (intentional —
     Marie fills these post-upload, e.g. Sorafenib refs, magnifications,
     wavelengths, STORM parentage) and surface unexpected sentinels.

CLI:
  python3 scripts/qa_flat_sheets.py [path/to/upload.xlsx]
"""

import os
import sys
import json
from pathlib import Path
from collections import defaultdict, Counter

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _config import config_from_args, plugin_context  # noqa: E402
from _config import ProjectRootError  # noqa: E402

# Sentinel substrings we treat as "deliberately deferred" — Marie fills
# these post-upload. Counted, not flagged as blockers.
EXPECTED_PLACEHOLDER_MARKERS = (
    "*** PLACEHOLDER",
    "***PLACEHOLDER",
)

# Surprise sentinels that almost always mean a build script forgot to
# fill something in. Surfaced separately.
SURPRISE_SENTINELS = ("XXX", "TODO", "FIXME", "???", "TBD", "UNCONFIRMED")


def load_prev_uids(prev_metadata):
    """Return set of all UIDs already in the master baseline workbook."""
    if prev_metadata is None or not Path(prev_metadata).exists():
        print(f"  ! WARNING: no master baseline workbook — parent-resolvability "
              f"will only check intra-upload UIDs")
        return set()
    wb = openpyxl.load_workbook(prev_metadata, data_only=True, read_only=True)
    uids = set()
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            continue
        try:
            uid_i = list(headers).index("UID")
        except ValueError:
            continue
        for r in rows:
            if r and r[uid_i]:
                uids.add(str(r[uid_i]).strip())
    wb.close()
    return uids


def load_sampletype_schemas(context_dir):
    """{short_code: {required: set, name: str}} from the sample type catalog.

    A project-local context/sampletypes_db.json wins over the plugin's bundled
    copy, so a project can pin a vintage.
    """
    local = Path(context_dir) / "sampletypes_db.json"
    path = local if local.is_file() else plugin_context("sampletypes_db.json")
    types = json.loads(Path(path).read_text())
    out = {}
    for t in types:
        code = t.get("SampleType")
        req = t.get("Required Metadata") or ""
        req_set = {x.strip() for x in req.split(",") if x.strip()}
        out[code] = {"required": req_set, "name": t.get("Name", "")}
    return out


def main(upload_path, cfg):
    prev_metadata = cfg.master_workbook
    expected_counts = cfg.expected_counts
    always_root = cfg.always_root or {"CEL", "MDL"}
    expected_total = sum(expected_counts.values()) if expected_counts else None
    print(f"Project root:    {cfg.root}")
    print(f"Upload file:     {upload_path}")
    print(f"Master baseline: {prev_metadata or '(none)'}")
    print()

    prev_uids = load_prev_uids(prev_metadata)
    schemas = load_sampletype_schemas(cfg.context)
    print(f"Master baseline: {len(prev_uids)} existing UIDs")
    print(f"NExtSEEK catalog: {len(schemas)} sample types")
    print()

    wb = openpyxl.load_workbook(upload_path, data_only=True, read_only=True)
    if "Samples" not in wb.sheetnames:
        print(f"✗ No 'Samples' sheet in {upload_path}")
        sys.exit(2)
    ws = wb["Samples"]
    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    try:
        uid_idx = headers.index("uid")
        st_idx = headers.index("sampletype")
        jm_idx = headers.index("json_metadata")
        parent_idx = headers.index("parent")
    except ValueError as e:
        print(f"✗ Required column missing: {e}")
        sys.exit(2)

    rows_per_uid = {}              # uid -> (row_n, sampletype)
    parents_referenced = []        # (row_n, uid, sampletype, parent_string)
    sampletype_counts = Counter()
    placeholder_counter = Counter()
    issues = defaultdict(list)
    names_by_st = defaultdict(set)
    rows_seen = 0
    blank_parent_roots = 0
    blank_parent_nonroot = 0

    for row_n, r in enumerate(rows_iter, start=2):
        if r is None or all(c is None for c in r):
            continue
        rows_seen += 1
        uid = r[uid_idx]
        st = r[st_idx]
        jm_str = r[jm_idx]
        parent_cell = r[parent_idx]

        sampletype_counts[st] += 1

        # (1) UID uniqueness
        if not uid:
            issues["missing_uid"].append(f"row {row_n}")
        elif uid in rows_per_uid:
            other_row, other_st = rows_per_uid[uid]
            issues["duplicate_uid"].append(
                f"{uid} appears in row {row_n} AND row {other_row} "
                f"(sampletype {other_st})")
        else:
            rows_per_uid[uid] = (row_n, st)

        # (1b) Collision with the LIVE DB (stamp-collision second net).
        # A NEW upload UID that already exists in the master baseline will
        # UPDATE (overwrite) that record on upload rather than insert — the
        # failure mode behind the 190220WHI/190219WHI incidents.
        if uid and uid in prev_uids:
            issues["uid_exists_in_db"].append(uid)

        # (2) Sampletype validity
        if st not in schemas:
            issues["unknown_sampletype"].append(f"row {row_n}: {st!r}")

        # (4) json_metadata JSON parse
        jm = {}
        if jm_str:
            try:
                jm = json.loads(jm_str)
            except json.JSONDecodeError as e:
                issues["bad_json"].append(f"row {row_n} ({uid}): {e}")
        elif st not in always_root:
            # Non-root rows with no json_metadata is suspicious
            issues["missing_json_metadata"].append(f"row {row_n} ({uid}, {st})")

        # (3) Parent: prefer the dedicated 'parent' column (denormalized),
        # fall back to JSON metadata 'Parent' key.
        parent = parent_cell or jm.get("Parent", "")
        if parent and str(parent).strip():
            parents_referenced.append((row_n, uid, st, str(parent)))
        else:
            if st in always_root:
                blank_parent_roots += 1
            else:
                blank_parent_nonroot += 1
                issues["blank_parent_nonroot"].append(
                    f"row {row_n} ({uid}, sampletype {st}) has no Parent")

        # (5) Required-metadata coverage (informational)
        req = schemas.get(st, {}).get("required", set())
        for field in req:
            if field == "UID":
                if not uid:
                    issues["missing_required"].append(
                        f"row {row_n}: missing UID")
            elif not jm.get(field):
                issues[f"missing_required:{st}:{field}"].append(
                    uid or f"row {row_n}")

        # (6) Name uniqueness within sampletype
        name = jm.get("Name") if isinstance(jm, dict) else None
        if name:
            if name in names_by_st[st]:
                issues["duplicate_name"].append(
                    f"sampletype {st}, name {name!r} reseen in row {row_n}")
            names_by_st[st].add(name)

        # (7) Placeholder sniff
        if jm_str:
            for marker in EXPECTED_PLACEHOLDER_MARKERS:
                if marker in jm_str:
                    placeholder_counter[marker] += jm_str.count(marker)
            for surprise in SURPRISE_SENTINELS:
                if surprise in jm_str:
                    placeholder_counter[f"surprise:{surprise}"] += 1

    wb.close()

    # (3) Validate Parent UID references against intra-upload UIDs ∪ prev MetNet
    all_uids = set(rows_per_uid)
    parent_resolution = {"intra": 0, "metnet": 0, "missing": 0}
    for row_n, uid, st, parent in parents_referenced:
        # Parent can be semicolon-joined for multi-parent OOC rows
        for p in [x.strip() for x in parent.split(";") if x.strip()]:
            if any(m in p for m in EXPECTED_PLACEHOLDER_MARKERS):
                continue
            if p in all_uids:
                parent_resolution["intra"] += 1
            elif p in prev_uids:
                parent_resolution["metnet"] += 1
            else:
                parent_resolution["missing"] += 1
                issues["parent_uid_not_found"].append(
                    f"row {row_n} {uid} ({st}) → Parent {p!r} "
                    f"not in upload or MetNet baseline")

    # ─── Report ────────────────────────────────────────────────────────────
    print("─" * 60)
    print("ROW COUNTS")
    print("─" * 60)
    print(f"Rows scanned: {rows_seen}")
    print(f"Unique new UIDs: {len(rows_per_uid)}")
    print()
    print("Per-sample-type:")
    all_st = sorted(set(sampletype_counts) | set(expected_counts))
    expected_ok = True
    for st in all_st:
        actual = sampletype_counts.get(st, 0)
        expected = expected_counts.get(st)
        if expected is None:
            marker = "" if not expected_counts else "  (unexpected sampletype!)"
            if expected_counts:
                expected_ok = False
            print(f"  {st:>6}: {actual:>5}{marker}")
        elif actual != expected:
            print(f"  {st:>6}: {actual:>5}  ✗ expected {expected}")
            expected_ok = False
        else:
            print(f"  {st:>6}: {actual:>5}  ✓")
    total_line = f"  {'TOTAL':>6}: {sum(sampletype_counts.values()):>5}"
    if expected_total is not None:
        total_line += f"  (expected {expected_total})"
    print(total_line)
    print()

    print("─" * 60)
    print("PARENT RESOLVABILITY")
    print("─" * 60)
    print(f"Parent references checked: {len(parents_referenced)}")
    print(f"  resolved intra-upload:    {parent_resolution['intra']}")
    print(f"  resolved via MetNet base: {parent_resolution['metnet']}")
    print(f"  UNRESOLVED:               {parent_resolution['missing']}")
    print(f"Rows with intentionally-blank Parent (root types): "
          f"{blank_parent_roots}")
    print()

    print("─" * 60)
    print("PLACEHOLDERS / SENTINELS")
    print("─" * 60)
    expected_placeholder_total = sum(v for k, v in placeholder_counter.items()
                                     if not k.startswith("surprise:"))
    surprise_total = sum(v for k, v in placeholder_counter.items()
                         if k.startswith("surprise:"))
    print(f"Expected placeholder occurrences (intentional, "
          f"Marie fills post-upload): {expected_placeholder_total}")
    for marker in EXPECTED_PLACEHOLDER_MARKERS:
        n = placeholder_counter.get(marker, 0)
        if n:
            print(f"  '{marker}': {n} occurrences")
    if surprise_total:
        print(f"  ⚠ UNEXPECTED sentinels: {surprise_total}")
        for k, v in placeholder_counter.items():
            if k.startswith("surprise:"):
                print(f"    '{k.split(':',1)[1]}': {v} rows")
    else:
        print("  (no unexpected sentinels)")
    print()

    # Issues
    print("─" * 60)
    print("ISSUES")
    print("─" * 60)
    blocker_count = 0

    # UID-vs-DB collision net. A NEW upload UID already present in the DB
    # baseline overwrites that record on upload — a stamp collision unless it is
    # a deliberate update/restore (then re-run with QA_ALLOW_DB_UPDATES=1).
    allow_updates = os.environ.get("QA_ALLOW_DB_UPDATES") == "1"
    collided = issues.get("uid_exists_in_db", [])
    if collided:
        tag = "INFO — updates acknowledged" if allow_updates else "BLOCKER"
        print(f"\n  [{tag}] {len(collided)} new UID(s) already exist in the DB baseline")
        print("    → on upload these UPDATE (overwrite) existing rows, not insert.")
        print("    → fresh batch? STAMP COLLISION with another study — re-mint under a")
        print("      free <YYMMDD><LAB> stamp (see scripts/stamp_guard.py).")
        print("    → intentional update/restore? re-run with QA_ALLOW_DB_UPDATES=1.")
        for u in collided[:10]:
            print(f"      - {u}")
        if len(collided) > 10:
            print(f"      ... and {len(collided) - 10} more")
        if not allow_updates:
            blocker_count += len(collided)

    blocker_keys = [
        "duplicate_uid", "missing_uid", "unknown_sampletype", "bad_json",
        "missing_json_metadata", "parent_uid_not_found", "duplicate_name",
        "blank_parent_nonroot",
    ]
    for k in blocker_keys:
        if k in issues:
            print(f"\n  [BLOCKER] {k} ({len(issues[k])}):")
            for line in issues[k][:10]:
                print(f"    - {line}")
            if len(issues[k]) > 10:
                print(f"    ... and {len(issues[k]) - 10} more")
            blocker_count += len(issues[k])

    # Required-field gap summary
    print(f"\n  [INFO] missing required-metadata fields per (sampletype, field):")
    req_groups = sorted(k for k in issues if k.startswith("missing_required:"))
    if not req_groups:
        print("    (none)")
    for k in req_groups:
        _, st, field = k.split(":", 2)
        print(f"    - {st:6s} {field:25s}  {len(issues[k]):3d} rows blank")

    print()
    print("=" * 60)
    if blocker_count == 0 and expected_ok:
        print(f"✓ PASS — no blockers across {rows_seen} rows, "
              f"all per-sample-type counts match")
    elif blocker_count == 0 and not expected_ok:
        print(f"⚠ row-count mismatch but no blocker issues — see above")
    else:
        print(f"✗ {blocker_count} blocker-level issues — see above")
    print("=" * 60)

    return 0 if blocker_count == 0 else 1


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description=__doc__)
    # NOTE: the config-override flags are declared inline (rather than via
    # _config.add_config_args) so that /curate-qa's doc contract — which names
    # --upload, --master-baseline and --expected-counts — is verifiable against
    # this script's own text (tests/test_curate_commands_present.py scans each
    # script file, not its imports). Dest names match add_config_args() so
    # config_from_args() consumes them identically.
    parser.add_argument(
        "--project-root", type=Path, default=None,
        help="Curation project root (default: nearest ancestor with "
             ".dmac-curation.json, else cwd)",
    )
    parser.add_argument(
        "--master-baseline", type=Path, default=None,
        help="Master baseline workbook whose UIDs upload parents may resolve "
             "against (default: newest previous_metadata/*All*.xlsx)",
    )
    parser.add_argument(
        "--expected-counts", default=None,
        help="Per-sampletype row expectations, e.g. 'OOC=122,CEL=2' "
             "(default: from the project lockfile, if any)",
    )
    parser.add_argument(
        "--upload", type=Path, default=None,
        help="Consolidated flat-format xlsx to QA "
             "(default: the single Arm*.xlsx under <project>/assay_sheets)",
    )
    parser.add_argument(
        "upload_pos", nargs="?", default=None, metavar="upload",
        help="Positional alias for --upload (deprecated; pass one or the other, "
             "not both).",
    )
    args = parser.parse_args()

    try:
        cfg = config_from_args(args)
    except ProjectRootError as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(2)

    if args.upload is not None and args.upload_pos is not None:
        print("ERROR: pass the upload as --upload OR positionally, not both",
              file=sys.stderr)
        sys.exit(2)
    upload = args.upload if args.upload is not None else args.upload_pos

    if upload is None:
        candidates = sorted(
            p for p in cfg.assay_sheets.glob("*.xlsx")
            if "_" not in p.stem and not p.name.startswith("~")
        )
        if len(candidates) != 1:
            print(f"ERROR: pass --upload; found {len(candidates)} consolidated "
                  f"sheets in {cfg.assay_sheets}: "
                  f"{[p.name for p in candidates]}", file=sys.stderr)
            sys.exit(2)
        upload = candidates[0]
    upload = Path(upload)
    if not upload.is_absolute():
        upload = (cfg.root / upload).resolve()
    if not upload.is_file():
        print(f"ERROR: {upload} does not exist", file=sys.stderr)
        sys.exit(2)
    sys.exit(main(upload, cfg))
