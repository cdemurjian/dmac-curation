# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""The lazy, cwd-only field dictionary.

No pre-built dictionary ships with the plugin, and none is generated for all
1059 field names. Each run creates entries only for the fields it touched, in
the current working directory.

Why lazy: the plugin already has a three-copies-of-context problem
(sampletypes_db.json exists in three places at three vintages). Shipping another
data file that drifts would repeat it. Lazy ships no state - nothing to version,
refresh, or go stale.

Accepted consequence: enrichment does not accumulate across projects. Ontology
IRIs curated while bolstering D.VIA in one directory do not help the next
curator elsewhere. The read-only inputs are shared; the enrichment is not.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

DICTIONARY_NAME = "field_dictionary.json"
OUTPUT_SUBDIR = "schema"

# Values that are markers, not data. Never recorded as observed vocabulary.
_NON_VALUES = ("*** PLACEHOLDER", "***PLACEHOLDER")


def observe_values(workbooks: list[Path], fields: set[str]) -> dict[str, list[str]]:
    """Distinct real values for `fields`, across every sheet of every workbook.

    Real observed values beat guessed ones (SKILL.md hard rule 4: schema lies,
    workbook tells truth). Order is first-seen, deduped, so a report reads the
    way the data does.
    """
    out: dict[str, list[str]] = {}
    for path in workbooks:
        path = Path(path)
        if not path.is_file():
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in wb.sheetnames:
                rows = wb[sheet].iter_rows(values_only=True)
                header = next(rows, None)
                if not header:
                    continue
                wanted = {i: str(h).strip() for i, h in enumerate(header)
                          if h and str(h).strip() in fields}
                if not wanted:
                    continue
                for row in rows:
                    for i, name in wanted.items():
                        if i >= len(row):
                            continue
                        value = row[i]
                        if value is None:
                            continue
                        text = str(value).strip()
                        if not text or any(m in text for m in _NON_VALUES):
                            continue
                        bucket = out.setdefault(name, [])
                        if text not in bucket:
                            bucket.append(text)
        finally:
            wb.close()
    return out


def build_entry(name, usage, observed: list[str], *, description: str = "",
                datatype: str = "string", synonyms: list[str] | None = None,
                ontology: dict | None = None,
                extra_provenance: str = "") -> dict:
    """One dictionary entry.

    `ontology` is always emitted with ``"confirmed": false``, whatever the
    caller passed. Only a human flips it, and this function is not a human.
    In the MUS prototype `Strain` was bound to NCBITaxon_10090, which is wrong -
    NCBITaxon covers species, not laboratory strains like C57BL/6J - and it was
    plausible enough to pass unreviewed.
    """
    if ontology is not None:
        ontology = dict(ontology)
        ontology["confirmed"] = False

    n_usage = len(getattr(usage, "used_by", []) or [])
    n_obs = len(observed)
    parts = [f"{n_usage} existing usage{'s' if n_usage != 1 else ''}"]
    parts.append(f"{n_obs} observed value{'s' if n_obs != 1 else ''}")
    if extra_provenance:
        parts.append(extra_provenance)

    return {
        "description": description,
        "datatype": datatype,
        "used_by": list(getattr(usage, "used_by", []) or []),
        "observed_values": list(observed),
        "ontology": ontology,
        "synonyms": list(synonyms or []),
        "provenance": " + ".join(parts),
    }


def merge_dictionary(existing: dict, new: dict) -> dict:
    """Merge a run's entries into a dictionary already on disk.

    Rules that matter:
      - observed_values and used_by are unioned, preserving first-seen order
      - a human-confirmed ontology binding is NEVER downgraded by an automated run
      - a non-empty human description is never replaced by an empty one
    """
    merged = {k: dict(v) for k, v in existing.items()}
    for name, entry in new.items():
        if name not in merged:
            merged[name] = dict(entry)
            continue
        old = merged[name]
        for list_key in ("observed_values", "used_by", "synonyms"):
            combined = list(old.get(list_key) or [])
            for v in entry.get(list_key) or []:
                if v not in combined:
                    combined.append(v)
            old[list_key] = combined
        if entry.get("description"):
            old["description"] = entry["description"]
        if entry.get("datatype"):
            old["datatype"] = entry["datatype"]
        old_ont, new_ont = old.get("ontology"), entry.get("ontology")
        if old_ont and old_ont.get("confirmed"):
            pass  # a human confirmed it; leave it alone
        elif new_ont:
            old["ontology"] = new_ont
        if entry.get("provenance"):
            old["provenance"] = entry["provenance"]
    return merged


def dictionary_path(root: Path) -> Path:
    return Path(root) / OUTPUT_SUBDIR / DICTIONARY_NAME


def load_dictionary(root: Path) -> dict:
    p = dictionary_path(root)
    if not p.is_file():
        return {}
    return json.loads(p.read_text())


def save_dictionary(root: Path, doc: dict) -> Path:
    p = dictionary_path(root)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(doc, indent=2, sort_keys=True) + "\n")
    return p
