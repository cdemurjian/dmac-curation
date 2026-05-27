#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
# Lifted from lee/scripts/apply_geo_accessions.py; generalized for dmac-curation plugin.
# Original contained hardcoded GSM/GSE accessions for a specific submission;
# replaced with CSV-driven approach. See TODO(v0.2) notes below.
"""Patch assay sheets with GEO accession URLs after GSE/GSM assignment.

Reads a GEO accession CSV (two columns: sample_id, gsm_accession) and patches:
  - D.SEQ upload sheet:   Accession = GSM, Link_PrimaryData = GSM URL
  - A.GEX upload sheet:   Link_PrimaryData = GSE URL (series-level, all rows)
  - A.SPTX upload sheet:  Link_PrimaryData = per-sample GSM URL

The CSV format (no header required; two tab- or whitespace-separated columns):
  GSM9751823    sample_title_ending_in_D123456
  ...

Sample IDs are extracted from each row's last _D###### token (bulk) or
D##-#### token (spatial). Rows whose sample type column starts with the
--uid-prefix are matched.

Usage:
  uv run scripts/apply_geo_accessions.py --gse-bulk GSE000001 --gsm-csv bulk.csv
  uv run scripts/apply_geo_accessions.py --gse-bulk GSE000001 --gsm-csv bulk.csv --write
"""
from __future__ import annotations

import argparse
import csv
import re
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SHEETS = ROOT / "assay_sheets"

GSM_URL = "https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc={}"

# TODO(v0.2): accept --sheets-dir to override SHEETS path


def parse_gsm_csv(csv_path: Path) -> dict[str, str]:
    """Parse a whitespace-delimited GSM roster into {sample_d_id: gsm_accession}.

    Each line: GSM<digits> <sample_title_ending_in_D######>
    The D-token is the last underscore-separated field or bracketed (D22-####/D23-####).
    """
    rx_bulk = re.compile(r"_(D\d+)\s*$")
    rx_sptx = re.compile(r"\((D\d{2}-\d{4})\)")
    m: dict[str, str] = {}
    with csv_path.open() as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split()
            if len(parts) < 2:
                continue
            gsm = parts[0]
            title = parts[1]
            hit = rx_bulk.search(title) or rx_sptx.search(title)
            if not hit:
                print(f"WARNING: could not extract D-id from {line!r}", file=sys.stderr)
                continue
            d_id = hit.group(1)
            if d_id in m:
                print(f"WARNING: duplicate D-id {d_id} ({m[d_id]} vs {gsm})", file=sys.stderr)
            m[d_id] = gsm
    return m


def extract_d_id(s: str | None, rx: re.Pattern) -> str | None:
    if not s:
        return None
    hit = rx.search(str(s))
    return hit.group(0) if hit else None


def patch_dseq(path: Path, gsm_map: dict[str, str], write: bool) -> None:
    print(f"\n=== {path.name} ===")
    wb = load_workbook(path)
    try:
        ws = wb.active
        rx = re.compile(r"D\d+")
        miss: list[tuple[int, str]] = []
        writes = 0
        for row in ws.iter_rows(min_row=2):
            uid = row[0].value
            if not uid or not str(uid).startswith("D.SEQ-"):
                continue
            fp = row[4].value  # File_PrimaryData (col E, 0-indexed=4)
            d = extract_d_id(fp, rx)
            if not d:
                miss.append((row[0].row, str(fp)))
                continue
            gsm = gsm_map.get(d)
            if not gsm:
                miss.append((row[0].row, f"{fp} (D={d} not in bulk roster)"))
                continue
            url = GSM_URL.format(gsm)
            if row[5].value != url:
                row[5].value = url  # Link_PrimaryData (col F)
                writes += 1
            if row[16].value != gsm:
                row[16].value = gsm  # Accession (col Q)
                writes += 1
        print(f"  rows to write: {writes}   unmapped: {len(miss)}")
        for r, why in miss[:8]:
            print(f"    !! row {r}: {why}")
        if miss[8:]:
            print(f"    ... +{len(miss) - 8} more")
        if write and not miss:
            shutil.copy(path, path.with_suffix(path.suffix + ".bak"))
            wb.save(path)
            print(f"  saved (.bak created)")
        elif write and miss:
            print("  refusing to save while unmapped rows exist")
    finally:
        wb.close()


def patch_agex(path: Path, gse_url: str, write: bool) -> None:
    print(f"\n=== {path.name} ===")
    wb = load_workbook(path)
    try:
        ws = wb.active
        writes = 0
        for row in ws.iter_rows(min_row=2):
            uid = row[0].value
            if not uid or not str(uid).startswith("A.GEX-"):
                continue
            if row[6].value != gse_url:
                row[6].value = gse_url  # Link_PrimaryData (col G)
                writes += 1
        gse_id = gse_url.split("acc=")[-1] if "acc=" in gse_url else gse_url
        print(f"  rows to write: {writes}   (all map to {gse_id})")
        if write:
            shutil.copy(path, path.with_suffix(path.suffix + ".bak"))
            wb.save(path)
            print(f"  saved (.bak created)")
    finally:
        wb.close()


def patch_asptx(path: Path, gsm_map: dict[str, str], write: bool) -> None:
    print(f"\n=== {path.name} ===")
    wb = load_workbook(path)
    try:
        ws = wb.active
        rx = re.compile(r"D\d{2}-\d{4}")
        miss: list[tuple[int, str]] = []
        writes = 0
        for row in ws.iter_rows(min_row=2):
            uid = row[0].value
            if not uid or not str(uid).startswith("A.SPTX-"):
                continue
            name = row[9].value  # Name column (col J, 0-indexed=9)
            d = extract_d_id(name, rx)
            if not d:
                miss.append((row[0].row, str(name)))
                continue
            gsm = gsm_map.get(d)
            if not gsm:
                miss.append((row[0].row, f"{name} (D={d} not in spatial roster)"))
                continue
            url = GSM_URL.format(gsm)
            if row[5].value != url:
                row[5].value = url  # Link_PrimaryData (col F)
                writes += 1
        print(f"  rows to write: {writes}   unmapped: {len(miss)}")
        for r, why in miss:
            print(f"    !! row {r}: {why}")
        if write and not miss:
            shutil.copy(path, path.with_suffix(path.suffix + ".bak"))
            wb.save(path)
            print(f"  saved (.bak created)")
        elif write and miss:
            print("  refusing to save while unmapped rows exist")
    finally:
        wb.close()


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--write", action="store_true", help="patch in place (creates .bak); default is dry-run")
    ap.add_argument(
        "--gse-bulk",
        required=True,
        metavar="GSE_ID",
        help="GEO series accession for bulk RNA-seq (e.g. GSE332646)",
    )
    ap.add_argument(
        "--gsm-csv",
        type=Path,
        required=True,
        metavar="CSV",
        help="Whitespace-delimited file: GSM<id>  <sample_title_with_D-token>",
    )
    ap.add_argument(
        "--gse-sptx",
        metavar="GSE_ID",
        default=None,
        help="GEO series accession for spatial data (optional; if omitted, spatial patch is skipped)",
    )
    ap.add_argument(
        "--sptx-gsm-csv",
        type=Path,
        default=None,
        metavar="CSV",
        help="GSM roster for spatial samples (uses --gsm-csv if omitted)",
    )
    ap.add_argument(
        "--sheets-dir",
        type=Path,
        default=SHEETS,
        metavar="DIR",
        help="Directory containing upload sheets (default: <project>/assay_sheets/)",
    )
    args = ap.parse_args()

    sheets = args.sheets_dir
    gse_bulk_url = GSM_URL.format(args.gse_bulk)

    gsm_map = parse_gsm_csv(args.gsm_csv)
    print(f"Bulk GSM roster: {len(gsm_map)} entries  ({args.gse_bulk})")

    sptx_map: dict[str, str] = {}
    if args.sptx_gsm_csv:
        sptx_map = parse_gsm_csv(args.sptx_gsm_csv)
    elif args.gse_sptx:
        # Fall back to bulk CSV for spatial tokens (D##-#### pattern)
        sptx_map = {k: v for k, v in gsm_map.items() if re.match(r"D\d{2}-\d{4}", k)}

    dseq_path = sheets / "D.SEQ-upload-new.xlsx"
    agex_path = sheets / "A.GEX-upload-new.xlsx"
    asptx_path = sheets / "A.SPTX-upload-new.xlsx"

    if dseq_path.exists():
        patch_dseq(dseq_path, gsm_map, args.write)
    else:
        print(f"\nWARNING: {dseq_path.name} not found — skipping D.SEQ patch")

    if agex_path.exists():
        patch_agex(agex_path, gse_bulk_url, args.write)
    else:
        print(f"\nWARNING: {agex_path.name} not found — skipping A.GEX patch")

    if args.gse_sptx:
        if asptx_path.exists():
            patch_asptx(asptx_path, sptx_map, args.write)
        else:
            print(f"\nWARNING: {asptx_path.name} not found — skipping A.SPTX patch")

    if not args.write:
        print("\n[dry run] re-run with --write to save")


if __name__ == "__main__":
    main()
