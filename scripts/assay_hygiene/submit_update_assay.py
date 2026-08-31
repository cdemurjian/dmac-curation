#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = ["requests>=2.31", "openpyxl>=3.1"]
# ///
"""Submit an UPDATE_ASSAY workbook to NExtSEEK, refusing anything that could delete.

PROVENANCE, BECAUSE THIS IS NOT NEW CODE. This is the submitter that performed
RUN1's registration write on 2026-08-27/28: 25,565 rows across 13 chunks,
verified afterwards as a bijection against the intended set -- every intended
row present, zero rows written that were not intended, zero duplicate pairs.

IT WAS ALMOST LOST. It lived in a session scratchpad and on the production box
at `~/ah_run2/`, never in this repository, so the plugin documented a write
stage whose actual tooling existed nowhere a later run could find it. RUN2
rediscovered that the hard way. It is vendored here, unchanged in logic, so the
next run inherits the gates rather than re-deriving them.

THE UID-UNIQUENESS GATE IS THE EXPENSIVE LESSON. Chunk 06 died mid-write at row
1221, having committed 1,220 rows, because four uids each matched TWO rows in
`samples`. `__retrieveSampleByUID` returns a record only when `len(records)==1`,
so a duplicated uuid resolves to None indistinguishably from a missing one, and
`getSampleID`'s None then reaches `if sample_id>0:` and raises TypeError --
500ing the whole batch with no feedback file written. The preflight that missed
it asked "does this uid EXIST" with a JOIN; the code asks "does exactly ONE row
have it". Those agree everywhere except on duplicates, which is the only case
that can hurt. Production carries duplicate-uuid samples and `samples.uuid` has
no unique constraint, so this is standing, not transient.

WHY A SHIM AND NOT curl. `POST /seek/sampleupload/` writes `assay_assets` one row
at a time with NO TRANSACTION, and the endpoint's own success report is not
trustworthy: `DBtable.storeOneRecord` (dmac/dbtable.py:109) sets `status = 1` and
never updates it from the DB call in either write branch, so a hard DB failure
comes back as success and the feedback workbook prints "successful:" for rows
that never wrote. Everything below therefore treats the response as a hint and
the database as the only receipt.

WHAT MAKES THE PAYLOAD SAFE. `updateSample_assay_asset`
(seek/dbtable_assay_assets.py:117) reaches `deleteOneRecord` only when BOTH
Current columns parse as int AND a New column does not. With both Current cells
blank, `int('')` raises, `id = -1`, and the delete branch behind `if id>0` is
structurally unreachable. The preflight below asserts that property of the actual
file rather than trusting whoever generated it.

DRY RUN BY DEFAULT. Nothing is sent without --confirm.
"""
from __future__ import annotations

import argparse
import csv
import getpass
import os
import re
import subprocess
import sys
from collections import Counter

import openpyxl
import requests

HEADERS = ["Sample UID", "Current Assay ID", "Current Assay Direction",
           "New Assay ID", "New Assay Direction"]
TOKEN_PAGE = "/seek/samples/upload/"
UPLOAD_PATH = "/seek/sampleupload/"


def _parses_int(value) -> bool:
    try:
        int(value)
        return True
    except (TypeError, ValueError):
        return False


def preflight(path: str, manifest_path: str | None,
              verify_uids: bool = True) -> tuple[int, Counter]:
    """Refuse the file unless every row is provably incapable of deleting."""
    book = openpyxl.load_workbook(path, data_only=True)
    problems: list[str] = []

    names = [n.upper().strip() for n in book.sheetnames]
    if "UPDATE_ASSAY" not in names:
        problems.append(f"no UPDATE_ASSAY sheet (found {book.sheetnames})")
    if "UPDATE" in names:
        problems.append("a sheet named UPDATE exists and would HIJACK the "
                        "dispatch into the metadata-update path "
                        "(seek/dbtable_sample.py:1663 is tested first)")
    if problems:
        raise SystemExit("REFUSED:\n  " + "\n  ".join(problems))

    sheet = book[book.sheetnames[names.index("UPDATE_ASSAY")]]
    header = [c.value for c in sheet[1]]
    if header != HEADERS:
        raise SystemExit(f"REFUSED: headers are case- and whitespace-sensitive "
                         f"(read verbatim as dict keys).\n  want {HEADERS}\n"
                         f"  got  {header}")

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise SystemExit("REFUSED: no data rows")

    # --- PROJECT CONSISTENCY -------------------------------------------------
    # SEEK assay ids are PER PROJECT. A registration must land on the assay
    # belonging to the SAMPLE'S OWN project, never merely on an assay with the
    # right title or the right internal id. The 2026-08-26 audit found 578 of
    # 26,188 rows targeting another project's assay -- all from the design's
    # "neighbour's own SEEK record" rule, where the lineage neighbour lived in
    # a different project. That is unrecoverable once written: the sample joins
    # a project it does not belong to.
    #
    # The check cannot be done from the workbook alone -- it needs each
    # sample's project set and each assay's project, which live in the extract.
    # So the builder emits MANIFEST.csv, gate-checked at build time, and this
    # asserts the sheet is exactly a subset of it. A regenerated or hand-edited
    # sheet that reintroduces a cross-project row fails here.
    if manifest_path:
        allowed = set()
        with open(manifest_path, newline="") as handle:
            for record in csv.DictReader(handle):
                if record["project_ok"].strip().lower() != "true":
                    raise SystemExit(
                        f"REFUSED: {manifest_path} contains a row flagged "
                        f"project_ok=false. The manifest itself is not clean.")
                allowed.add((record["uid"], str(int(float(record["assay_id"])))))
        unlisted = [(u, n) for u, _, _, n, _ in rows
                    if (str(u), str(n)) not in allowed]
        if unlisted:
            raise SystemExit(
                f"REFUSED: {len(unlisted)} row(s) are not in the gate-checked "
                f"manifest, so their project consistency is unproven. "
                f"First: {unlisted[:3]}")
    else:
        print("WARNING: no --manifest given. Project consistency is UNVERIFIED. "
              "SEEK assay ids are per-project and a cross-project write cannot "
              "be undone by re-running.", file=sys.stderr)

    # --- UID UNIQUENESS, checked LIVE against the database -------------------
    # `__retrieveSampleByUID` (seek/dbtable_sample.py) returns a record only
    # when `len(records) == 1`. TWO matching rows returns None, indistinguish-
    # ably from zero -- and `getSampleID`'s None then reaches `if sample_id>0:`
    # at :1564, which raises TypeError and 500s the WHOLE batch, leaving every
    # row before it committed. That is what killed chunk 06 on 2026-08-27 at
    # row 1221, after 1,220 rows had already landed.
    #
    # Production holds 15 duplicate-uuid samples, created 2026-03-11 by an
    # upload that inserted each twice. `samples.uuid` has no unique constraint.
    #
    # THIS IS A LIVE QUERY ON PURPOSE. The earlier preflight checked existence
    # with a JOIN, and a uid matching two rows still counts as "resolving" --
    # existence and uniqueness differ on precisely the case that matters. A
    # manifest column would be a snapshot and would go stale the same way.
    if verify_uids:
        uids = sorted({str(r[0]) for r in rows if isinstance(r[0], str)})
        values = ",".join(f"('{u}')" for u in uids)
        sql = (
            "CREATE TEMPORARY TABLE ah_uq (u VARCHAR(64) NOT NULL PRIMARY KEY) "
            "ENGINE=InnoDB;\n"
            f"INSERT IGNORE INTO ah_uq VALUES {values};\n"
            "SELECT a.u, COUNT(*) FROM ah_uq a "
            "JOIN seek_production.samples s ON s.uuid=a.u "
            "GROUP BY a.u HAVING COUNT(*) <> 1;\n"
            "SELECT CONCAT('__MISSING__', a.u) FROM ah_uq a "
            "LEFT JOIN seek_production.samples s ON s.uuid=a.u "
            "WHERE s.id IS NULL;\n")
        proc = subprocess.run(
            ["docker", "exec", "-i", "seek-mysql", "sh", "-c",
             'mysql -h"$MYSQL_HOST" -uroot -p"$MYSQL_ROOT_PASSWORD" '
             '-N -B seek_production'],
            input=sql, capture_output=True, text=True)
        if proc.returncode != 0:
            raise SystemExit(
                "REFUSED: could not verify UID uniqueness against the "
                f"database:\n  {proc.stderr.strip()[:400]}\n"
                "Re-run with --no-verify-uids only if you have checked by hand "
                "that every UID resolves to exactly one samples row.")
        offenders = [ln for ln in proc.stdout.splitlines()
                     if ln.strip() and "Using a password" not in ln]
        if offenders:
            raise SystemExit(
                f"REFUSED: {len(offenders)} UID(s) do not resolve to exactly "
                f"one samples row. Each one 500s the ENTIRE batch at "
                f"dbtable_sample.py:1564, leaving every earlier row committed "
                f"and no feedback file written:\n  "
                + "\n  ".join(offenders[:10]))
        print(f"  uid uniqueness      : VERIFIED live ({len(uids):,} uids, "
              f"all resolve to exactly 1 row)")
    else:
        print("WARNING: --no-verify-uids given. A UID matching zero OR TWO "
              "samples rows will 500 the whole batch mid-write.",
              file=sys.stderr)

    could_delete = would_noop = bad_uid = 0
    targets: Counter = Counter()
    for uid, cur_id, cur_dir, new_id, new_dir in rows:
        current_parses = _parses_int(cur_id) and _parses_int(cur_dir)
        new_parses = _parses_int(new_id) and _parses_int(new_dir)
        if current_parses and not new_parses:
            could_delete += 1
        if not new_parses:
            would_noop += 1
        if uid is None or not isinstance(uid, str) or not uid.strip():
            bad_uid += 1
        if new_parses:
            targets[int(new_id)] += 1

    if could_delete:
        problems.append(f"{could_delete} row(s) would DELETE an existing edge")
    if would_noop:
        problems.append(f"{would_noop} row(s) would silently no-op (a New cell "
                        f"does not parse as int)")
    if bad_uid:
        problems.append(f"{bad_uid} UID cell(s) are blank, non-string or "
                        f"whitespace-only. getSampleID returns None and "
                        f"`if sample_id>0` raises TypeError, which 500s the "
                        f"WHOLE run mid-write with rows already committed")
    if problems:
        raise SystemExit("REFUSED:\n  " + "\n  ".join(problems))
    return len(rows), targets


def submit(base: str, path: str, username: str, password: str,
           timeout: int) -> dict:
    session = requests.Session()
    page = session.get(base.rstrip("/") + TOKEN_PAGE, timeout=timeout)
    page.raise_for_status()

    token = session.cookies.get("csrftoken")
    if not token:
        found = re.search(r'name="csrfmiddlewaretoken"\s+value="([^"]+)"',
                          page.text)
        token = found.group(1) if found else None
    if not token:
        raise SystemExit("REFUSED: no CSRF token from " + TOKEN_PAGE +
                         " — are you pointed at the right host?")

    with open(path, "rb") as handle:
        response = session.post(
            base.rstrip("/") + UPLOAD_PATH,
            # getSeekLogin reads username/password straight off POST
            # (seek/seekdb.py:171-176). Not Django-logged-in, so
            # verifySuperUser is 0 and no creator id is required.
            data={"csrfmiddlewaretoken": token,
                  "username": username, "password": password},
            files={"excelfile_upload": (os.path.basename(path), handle,
                   "application/vnd.openxmlformats-officedocument."
                   "spreadsheetml.sheet")},
            headers={"X-CSRFToken": token, "Referer": base.rstrip("/") + TOKEN_PAGE},
            timeout=timeout)
    response.raise_for_status()
    try:
        return response.json()
    except ValueError:
        return {"_raw": response.text[:2000]}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--base-url", default="http://127.0.0.1:8000")
    parser.add_argument("--no-verify-uids", action="store_true",
                        help="skip the live check that every UID resolves to "
                             "exactly ONE samples row. Only for a host with no "
                             "database access; a duplicate uid 500s the batch.")
    parser.add_argument("--manifest", default=None,
                        help="MANIFEST.csv from the builder. Proves every row "
                             "targets an assay in the SAMPLE'S OWN project. "
                             "Omitting it is allowed but loudly unverified.")
    parser.add_argument("--username", required=True)
    parser.add_argument("--password-env", default="NEXTSEEK_PASSWORD",
                        help="env var holding the password; prompted if unset. "
                             "Never pass a password in argv -- it is visible "
                             "in ps to every user on the box.")
    parser.add_argument("--max-rows", type=int, default=250,
                        help="hard cap. Raise it deliberately; the default "
                             "exists so the full 26,188-row sheet cannot be "
                             "sent by reflex.")
    parser.add_argument("--timeout", type=int, default=1800)
    parser.add_argument("--confirm", action="store_true",
                        help="actually POST. Without this it is a dry run.")
    args = parser.parse_args()

    n_rows, targets = preflight(args.sheet, args.manifest,
                                verify_uids=not args.no_verify_uids)
    print(f"PREFLIGHT PASSED  {args.sheet}")
    print(f"  rows                : {n_rows:,}")
    print(f"  distinct assays     : {len(targets)}")
    print(f"  rows able to delete : 0")
    print(f"  project-consistent  : "
          f"{'YES (manifest verified)' if args.manifest else 'UNVERIFIED'}")
    print(f"  rows that would noop: 0")
    print(f"  per-assay additions : "
          f"{dict(sorted(targets.items(), key=lambda kv: -kv[1]))}")

    if n_rows > args.max_rows:
        raise SystemExit(f"\nREFUSED: {n_rows:,} rows exceeds --max-rows "
                         f"{args.max_rows:,}.")

    if not args.confirm:
        print("\nDRY RUN — nothing sent. Re-run with --confirm to submit.")
        print("Before you do: capture the rollback handle and a backup.")
        print("  SELECT MAX(id) FROM assay_assets;")
        return

    password = os.environ.get(args.password_env) or getpass.getpass("SEEK password: ")
    print(f"\nPOSTing to {args.base_url.rstrip('/')}{UPLOAD_PATH} ...")
    result = submit(args.base_url, args.sheet, args.username, password,
                    args.timeout)
    print("\n--- response ---")
    for key, value in result.items():
        print(f"  {key}: {str(value)[:600]}")
    print("\nThe response and the feedback workbook are NOT a write receipt:")
    print("  DBtable.storeOneRecord returns status=1 even when the DB write")
    print("  failed. Verify against the database before believing any of it.")


if __name__ == "__main__":
    main()
