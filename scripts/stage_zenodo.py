#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
# Lifted from lee/scripts/stage_zenodo.py; generalized for dmac-curation plugin.
"""Stage curated non-image files into per-figure data-type folders for Zenodo zipping.

Walks files/Figure {N}/ and files/Source Data/. For each curated, non-image file:
  - Determines its target figure N
  - Moves it to files/Figure {N}/Figure{N}_{SampleType}/{filename}

Image sample types (D.IMG, A.IMG, SLD, A.SPTX) are excluded — those go to OMERO/GEO.
files/Figures/ is skipped — it may have duplicate blot files already in Source Data/.

Usage:
  uv run scripts/stage_zenodo.py                                     # report what would happen (default)
  uv run scripts/stage_zenodo.py --write                             # actually move files
  uv run scripts/stage_zenodo.py --metadata-xlsx path/to/metadata.xlsx
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _config import add_config_args, config_from_args  # noqa: E402
from _config import ProjectRootError  # noqa: E402

IMAGE_TYPES = {"D.IMG", "A.IMG", "SLD", "A.SPTX"}

# Default figure assignment for shared sample types whose files don't live in a Figure N/ dir.
# TODO(v0.2): make this configurable per-project (e.g. via a JSON sidecar).
SHARED_FIGURE_DEFAULT: dict[str, int] = {}


def load_curation_index(meta: Path) -> dict[str, tuple[str, str]]:
    """Build {filename → (SampleType, UID)} from all sheets in the metadata workbook."""
    wb = load_workbook(meta, read_only=True, data_only=True)
    try:
        out: dict[str, tuple[str, str]] = {}
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            hdr = None
            for r in ws.iter_rows(values_only=True):
                if hdr is None:
                    hdr = list(r)
                    i_uid = hdr.index("UID") if "UID" in hdr else None
                    i_file = hdr.index("File_PrimaryData") if "File_PrimaryData" in hdr else None
                    if i_uid is None or i_file is None:
                        break
                    continue
                if r[i_file]:
                    fn = str(r[i_file]).strip()
                    if fn not in out:
                        out[fn] = (sheet, str(r[i_uid]))
        return out
    finally:
        wb.close()


def figure_for_path(p: Path, files_dir: Path) -> int | None:
    """Extract figure number from a path under files/Figure N/."""
    parts = p.relative_to(files_dir).parts
    if not parts:
        return None
    if parts[0].startswith("Figure ") and len(parts[0].split()) > 1:
        try:
            return int(parts[0].split()[1])
        except ValueError:
            return None
    return None


def already_staged(p: Path) -> bool:
    """True if file is already inside a FigureN_SampleType staging subfolder."""
    return bool(re.search(r"/Figure\d+_[A-Z.]+/", str(p)))


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    add_config_args(ap)
    ap.add_argument(
        "--write", action="store_true",
        help="Move files into staging folders; default is dry-run.",
    )
    ap.add_argument(
        "--metadata-xlsx",
        metavar="XLSX",
        help="Path to All-Metadata workbook (default: newest previous_metadata/*All*.xlsx)",
    )
    args = ap.parse_args()

    try:
        cfg = config_from_args(args)
    except ProjectRootError as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(2)

    FILES = cfg.files
    meta = Path(args.metadata_xlsx).resolve() if args.metadata_xlsx else cfg.master_workbook
    if meta is None or not meta.exists():
        print("ERROR: no metadata xlsx found. Pass --metadata-xlsx <path>.", file=sys.stderr)
        sys.exit(1)
    print(f"Project root: {cfg.root}")
    print(f"Loading curation index from {meta.name}...")
    idx = load_curation_index(meta)
    print(f"  {len(idx)} curated filenames\n")

    plan: list[tuple[Path, Path, str, str]] = []  # (src, dst, sample_type, uid)
    skipped_uncurated: list[Path] = []
    skipped_already_staged: list[Path] = []
    skipped_image: list[tuple[Path, str]] = []
    skipped_other: list[tuple[Path, str]] = []
    collisions: list[tuple[Path, Path]] = []

    # Walk only Figure N/ + Source Data/, skip files/Figures/
    for root in sorted(FILES.iterdir()):
        if not root.is_dir():
            continue
        if root.name == "Figures":
            continue
        for p in root.rglob("*"):
            if not p.is_file():
                continue
            if p.name.startswith("._") or p.name == ".DS_Store":
                continue
            if already_staged(p):
                skipped_already_staged.append(p)
                continue
            entry = idx.get(p.name)
            if not entry:
                skipped_uncurated.append(p)
                continue
            stype, uid = entry
            if stype in IMAGE_TYPES:
                skipped_image.append((p, stype))
                continue
            # Determine target figure
            fig = figure_for_path(p, FILES)
            if fig is None:
                # Shared / outside Figure N dir — use default mapping
                fig = SHARED_FIGURE_DEFAULT.get(stype)
                if fig is None:
                    skipped_other.append((p, f"no figure mapping for {stype}"))
                    continue
            dst_dir = FILES / f"Figure {fig}" / f"Figure{fig}_{stype}"
            dst = dst_dir / p.name
            if dst.exists() and dst != p:
                collisions.append((p, dst))
                continue
            plan.append((p, dst, stype, uid))

    # Group plan by (figure, sample_type) for the summary
    by_bucket: dict[tuple[int, str], list] = defaultdict(list)
    for src, dst, stype, uid in plan:
        m = re.search(r"Figure (\d+)/Figure\d+_", str(dst))
        fig = int(m.group(1)) if m else 0
        by_bucket[(fig, stype)].append((src, dst, uid))

    print("=== Plan: files to stage ===")
    for (fig, stype), items in sorted(by_bucket.items()):
        print(f"  Figure {fig} / {stype}: {len(items)} files")
        for src, dst, uid in items[:3]:
            rel = src.relative_to(FILES)
            print(f"    {uid:25s} {rel}")
        if len(items) > 3:
            print(f"    ... +{len(items)-3} more")

    print(f"\n=== Summary ===")
    print(f"  files to move:       {len(plan)}")
    print(f"  skipped (uncurated): {len(skipped_uncurated)}")
    print(f"  skipped (already staged): {len(skipped_already_staged)}")
    print(f"  skipped (image, → OMERO/GEO not Zenodo): {len(skipped_image)}")
    print(f"  skipped (other):     {len(skipped_other)}")
    print(f"  collisions:          {len(collisions)}")

    if skipped_image[:5]:
        print(f"\n  Image-type sample of skipped:")
        cnt = defaultdict(int)
        for _, t in skipped_image:
            cnt[t] += 1
        for t, n in sorted(cnt.items()):
            print(f"    {t}: {n} files")

    if skipped_other[:10]:
        print(f"\n  Other-skip examples:")
        for p, why in skipped_other[:10]:
            print(f"    {p.relative_to(FILES)}: {why}")

    if collisions[:5]:
        print(f"\n  Collisions (destination already exists, will NOT overwrite):")
        for src, dst in collisions[:10]:
            print(f"    {src.relative_to(FILES)} → {dst.relative_to(FILES)}")

    if not args.write:
        print("\nDry run — no files moved. Re-run with --write to execute.")
        return

    print(f"\nMoving {len(plan)} files...")
    moved = 0
    for src, dst, stype, uid in plan:
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(src), str(dst))
        moved += 1
    print(f"  Moved {moved} files into per-figure data-type folders.")


if __name__ == "__main__":
    main()
