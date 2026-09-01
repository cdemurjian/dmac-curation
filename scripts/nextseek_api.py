#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = ["requests>=2.31", "openpyxl>=3.1"]
# ///
"""
Reusable NExtSEEK API client for resolving assay titles → assay IDs per project.

NExtSEEK structure: projects → studies → assays. The set of assays in any
given project is project-specific, so the same assay name (e.g. "RNA Extraction")
maps to different `Internal Assay ID` values depending on the project.

Auth: HTTP Basic (spec § securitySchemes.basicAuth at YAML line ~8976). Read
credentials from `.env` at the repo root or directly from env vars:
  NEXTSEEK_USERNAME=...
  NEXTSEEK_PASSWORD=...
Or pass `--username` / `--password` on the CLI (less safe — appears in shell
history). Token auth is also supported via NEXTSEEK_TOKEN / --token if you
later switch.

CLI usage:
  # populate .env then run:
  python scripts/nextseek_api.py fetch-assays --project-id 10
    → writes context/assay_ids_cache.json keyed by assay title

Module usage:
  from scripts.nextseek_api import NExtSEEKClient
  client = NExtSEEKClient(username="...", password="...")
  id_map = client.fetch_assay_id_map(project_id=10)
  # → {"RNA Extraction": 61, "Tissue Collection": 74, ...}
"""

import argparse
import json
import os
import sys
import time
from pathlib import Path
from typing import Dict, Iterator, Optional, Tuple

import requests
from requests.auth import HTTPBasicAuth

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _config import add_config_args, config_from_args  # noqa: E402
from _config import ProjectRootError  # noqa: E402

# Kept only for the plugin-local .env fallback; never used for project paths.
_PLUGIN = Path(__file__).resolve().parent.parent
# nextseek.mit.edu serves `/nextseek_api/` (line 3462 of the spec self-references
# nextseek-dev.mit.edu as the schema host). fairdata.mit.edu is the SEEK web UI,
# not the API. Override via --base-url if pointing at dev or a different deployment.
DEFAULT_BASE_URL = "https://nextseek.mit.edu"


class NExtSEEKError(RuntimeError):
    """Wraps a NExtSEEK API failure with status + response body for debugging."""

    def __init__(self, status: int, url: str, body: str):
        super().__init__(f"NExtSEEK {status} on {url}\n{body[:1000]}")
        self.status = status
        self.url = url
        self.body = body


class NExtSEEKClient:
    """Thin client for the read endpoints we need to resolve assay IDs.

    Not a full SDK — just enough to:
      - GET /projects/{id}/ → traverse relationships.assays.data for assay IDs
      - GET /assays/ (paginated) → harvest (id, title) pairs
      - intersect the two to produce {title: id} for project-scoped assays

    Auth: pass (username, password) for HTTP Basic, OR token for Token auth.
    If both are passed, Basic wins.
    """

    def __init__(self, username: Optional[str] = None,
                 password: Optional[str] = None,
                 token: Optional[str] = None,
                 base_url: str = DEFAULT_BASE_URL,
                 timeout: float = 30.0):
        if not (username and password) and not token:
            raise ValueError("provide (username, password) or token")
        self.base_url = base_url.rstrip("/")
        self.timeout = timeout
        self.session = requests.Session()
        self.session.headers.update({
            "Accept": "application/json",
            "User-Agent": "dmac-curation/0.1.0",
        })
        if username and password:
            self.session.auth = HTTPBasicAuth(username, password)
            self.auth_mode = "basic"
        else:
            self.session.headers["Authorization"] = f"Token {token}"
            self.auth_mode = "token"

    # ── Low-level GET with consistent error surface ─────────────────────────

    def _get(self, path: str, params: Optional[dict] = None) -> dict:
        url = f"{self.base_url}/nextseek_api{path}"
        resp = self.session.get(url, params=params, timeout=self.timeout)
        if not resp.ok:
            raise NExtSEEKError(resp.status_code, url, resp.text)
        try:
            return resp.json()
        except ValueError:
            raise NExtSEEKError(resp.status_code, url,
                                f"Non-JSON response: {resp.text[:500]}")

    # ── Endpoint wrappers ───────────────────────────────────────────────────

    def export_project(self, project_id,
                       output_format: str = "xlsx") -> Tuple[bytes, str]:
        """GET /admin/project-export/{id}/ → (content_bytes, server_filename).

        The full project database export. ``output_format='xlsx'`` returns the
        master workbook (one sheet per sample type); ``'json'`` returns JSON
        bytes. This is the fresh DB pull the build stamp-guard checks against.
        Returns raw bytes (not parsed) so a binary xlsx passes through intact.
        """
        import re
        url = f"{self.base_url}/nextseek_api/admin/project-export/{project_id}/"
        resp = self.session.get(url, params={"output_format": output_format},
                                timeout=max(self.timeout, 300.0))
        if not resp.ok:
            raise NExtSEEKError(resp.status_code, url, resp.text[:500])
        cd = resp.headers.get("Content-Disposition", "")
        m = re.search(r'filename="?([^";]+)"?', cd)
        ext = "xlsx" if output_format == "xlsx" else "json"
        fname = m.group(1).strip() if m else f"project-{project_id}-export.{ext}"
        return resp.content, fname

    def get_project(self, project_id) -> dict:
        """GET /projects/{id}/ — returns the full JSON:API response dict."""
        return self._get(f"/projects/{project_id}/")

    def list_projects(self) -> list:
        """GET /projects/ → [{'id': int, 'title': str}, ...] (JSON:API normalized)."""
        doc = self._get("/projects/")
        data = doc.get("data", doc) if isinstance(doc, dict) else doc
        out = []
        for r in (data or []):
            attrs = r.get("attributes", r) if isinstance(r, dict) else {}
            pid = r.get("id")
            try:
                pid = int(pid)
            except (TypeError, ValueError):
                pass
            out.append({"id": pid, "title": attrs.get("title") or attrs.get("name") or ""})
        return out

    def get_sample_type(self, ident) -> dict:
        """GET /sample_types/{id-or-title}/ — full schema incl. sample_attributes."""
        return self._get(f"/sample_types/{ident}/")

    def patch_sample_type(self, type_id, sample_attributes: list) -> dict:
        """PATCH /sample_types/{id}/ — replace the sample_attributes array.

        DESTRUCTIVE IF MISUSED. The server treats `sample_attributes` as the
        COMPLETE list, not a delta: any existing attribute omitted from the
        array is dropped from the type. Callers must send every current
        attribute (each carrying its `id`) plus whatever is being added.
        `add_sample_type_attribute()` below is the safe wrapper — prefer it.

        Sample types are GLOBAL. A change here affects every project and every
        existing record of this type across NExtSEEK, not just the caller's.
        """
        csrf = self._prime_csrf()
        headers = {"Content-Type": "application/json"}
        if csrf:
            headers["X-CSRFToken"] = csrf
            headers["Referer"] = self.base_url
        url = f"{self.base_url}/nextseek_api/sample_types/{type_id}/"
        payload = {"data": {"id": str(type_id), "type": "sample_types",
                            "attributes": {"sample_attributes": sample_attributes}}}
        resp = self.session.patch(url, json=payload, headers=headers,
                                  timeout=self.timeout)
        if not resp.ok:
            raise NExtSEEKError(resp.status_code, url, resp.text)
        try:
            return resp.json()
        except ValueError:
            raise NExtSEEKError(resp.status_code, url,
                                f"Non-JSON response: {resp.text[:500]}")

    def _write(self, method: str, path: str, payload: dict) -> dict:
        """POST/PATCH JSON to /nextseek_api{path}, with CSRF primed.

        Shared by the attribute endpoints. Django enforces CSRF on unsafe
        methods regardless of auth mode, and the /nextseek_api/* views do not
        issue the cookie themselves -- see `_prime_csrf`.
        """
        csrf = self._prime_csrf()
        headers = {"Content-Type": "application/json"}
        if csrf:
            headers["X-CSRFToken"] = csrf
            headers["Referer"] = self.base_url
        url = f"{self.base_url}/nextseek_api{path}"
        resp = self.session.request(method, url, json=payload, headers=headers,
                                    timeout=self.timeout)
        if not resp.ok:
            raise NExtSEEKError(resp.status_code, url, resp.text)
        try:
            return resp.json()
        except ValueError:
            raise NExtSEEKError(resp.status_code, url,
                                f"Non-JSON response: {resp.text[:500]}")

    # ── Sample-type attributes ──────────────────────────────────────────────
    #
    # THE REAL WRITE PATH, live on production since 2026-08-31. It replaces
    # `scripts/sampletype_attr.py`, which drove the admin UI's own
    # `GET /seek/attribute/save/?records=<json>` because no REST write path
    # existed. That is no longer true, and the differences are not cosmetic:
    #
    #   * `dry_run` is planned SERVER-SIDE, so the preview is the server's own
    #     answer rather than the client's guess at it
    #   * errors are structured JSON (409 per-target documents carrying
    #     `submitted_identifier`, 422 naming the offending field) instead of an
    #     HTML page scraped for a success string
    #   * the server enforces title uniqueness and the single-title-attribute
    #     rule, so the client does not re-implement validation it cannot keep
    #     in step with
    #
    # Identifiers (`sample_type`, `sample_attribute_type`) accept a database
    # id, a numeric string, or the exact title.
    #
    # MUTATIONS REQUIRE A DJANGO SUPERUSER (is_superuser), which is NOT the
    # same population as a SEEK admin -- the two are not nested. Reads need
    # only a SEEK login. A SEEK admin who is not a Django superuser gets 403
    # `permission_denied` with "Superuser access required."

    ATTRIBUTES_PATH = "/attributes/"

    def list_attributes(self, sample_type=None) -> dict:
        """GET /attributes/ — reads, open to any SEEK-authenticated user."""
        params = {"sample_type": sample_type} if sample_type else None
        return self._get(self.ATTRIBUTES_PATH, params=params)

    @staticmethod
    def _targets(sample_type, attributes: list) -> dict:
        return {"targets": [{"sample_type": sample_type,
                             "attributes": attributes}]}

    def create_attributes(self, sample_type, attributes: list,
                          dry_run: bool = True) -> dict:
        """POST /attributes/batch-create/ — add attributes to a sample type.

        GLOBAL, SHARED-SCHEMA WRITE. Sample types are not project-scoped:
        adding a field changes that type for every project and every existing
        record of it across NExtSEEK.
        """
        payload = self._targets(sample_type, attributes) | {"dry_run": dry_run}
        return self._write("POST", f"{self.ATTRIBUTES_PATH}batch-create/",
                           payload)

    def patch_attributes(self, sample_type, attributes: list,
                         dry_run: bool = True) -> dict:
        """PATCH /attributes/batch-patch/ — amend existing attributes."""
        payload = self._targets(sample_type, attributes) | {"dry_run": dry_run}
        return self._write("PATCH", f"{self.ATTRIBUTES_PATH}batch-patch/",
                           payload)

    def delete_attributes(self, sample_type, attributes: list,
                          dry_run: bool = True) -> dict:
        """POST /attributes/batch-delete/ — remove attributes.

        DESTRUCTIVE AND GLOBAL. Deleting an attribute removes it from every
        record of the type. There is no undo through this API.
        """
        payload = self._targets(sample_type, attributes) | {"dry_run": dry_run}
        return self._write("POST", f"{self.ATTRIBUTES_PATH}batch-delete/",
                           payload)

    def _prime_csrf(self) -> Optional[str]:
        """GET /login/ to populate the csrftoken cookie, then return its value.

        Django enforces CSRF on POST regardless of auth method. The csrftoken
        cookie is issued only by template-rendering views like /login/ — the
        /nextseek_api/* endpoints don't set it. We GET /login/ (which returns
        200 + Set-Cookie: csrftoken=…), pull the token from the session jar,
        then echo it on the POST as X-CSRFToken. The Referer header is also
        required by Django CSRF over HTTPS.
        """
        try:
            self.session.get(f"{self.base_url}/login/", timeout=self.timeout)
        except requests.RequestException:
            pass  # non-fatal; POST will fail with 403 if cookie missing
        return self.session.cookies.get("csrftoken")

    def validate_batch_upload(self, file_path: Path, project_id,
                              checks: str = "structure") -> dict:
        """POST /batch-upload/validate/ — dry-run validation, no side effects.

        Returns ValidationResult dict with: valid (bool), summary (str),
        totals (dict), errors (list), warnings (dict), checks_run (list),
        checks_skipped (list). See spec line 640.

        checks: comma-separated subset of 'structure,name_check,dag'.
                'structure' is fastest (CONVERT + json_metadata attr check).
                'dag' builds parent/child graph, reports orphans + cycles.
                'name_check' verifies sample Name doesn't already exist in DB.
        """
        # Django CSRF on POST: prime the cookie + echo it in X-CSRFToken header.
        csrf = self._prime_csrf()
        headers = {}
        if csrf:
            headers["X-CSRFToken"] = csrf
            headers["Referer"] = self.base_url  # Django CSRF also checks Referer for HTTPS

        url = f"{self.base_url}/nextseek_api/batch-upload/validate/"
        with open(file_path, "rb") as fh:
            files = {
                "file": (file_path.name, fh,
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            }
            data = {
                "project_id": str(project_id),
                "checks": checks,
            }
            resp = self.session.post(url, files=files, data=data,
                                     headers=headers, timeout=self.timeout)
        if not resp.ok:
            raise NExtSEEKError(resp.status_code, url, resp.text)
        try:
            return resp.json()
        except ValueError:
            raise NExtSEEKError(resp.status_code, url,
                                f"Non-JSON response: {resp.text[:500]}")

    def list_assays_paginated(self, page_size: int = 100,
                              verbose: bool = True) -> Iterator[Tuple[str, str]]:
        """GET /assays/ — yields (id, title) tuples across all pages.

        Termination rule: stop when `links.next` is null/missing. The MIT
        NExtSEEK deployment ignores `page[size]` and returns the full result
        set in one response (with next=null), so this typically makes one
        request. We DON'T fall back to `len(records) < page_size` — that
        heuristic causes an infinite loop when the server ignores pagination.
        """
        page = 1
        while True:
            t0 = time.monotonic()
            body = self._get("/assays/", params={
                "page[number]": page,
                "page[size]": page_size,
            })
            envelope = body[0] if isinstance(body, list) and body else body
            records = envelope.get("data") or []
            elapsed = time.monotonic() - t0
            if verbose:
                print(f"    GET /assays/ page {page}: {len(records)} records "
                      f"in {elapsed:.2f}s", file=sys.stderr)

            for rec in records:
                aid = str(rec.get("id"))
                title = (rec.get("attributes") or {}).get("title", "")
                if aid and title:
                    yield aid, title

            # Trust the server's pagination link as the sole termination signal.
            next_link = (envelope.get("links") or {}).get("next")
            if not next_link:
                break
            page += 1
            time.sleep(0.05)

    # ── High-level: title → id map scoped to one project ────────────────────

    def fetch_assay_id_map(self, project_id, verbose: bool = True) -> Dict[str, int]:
        """Return {assay_title: assay_id} for assays linked to a project.

        Strategy:
          1. GET project → harvest its relationships.assays.data IDs
          2. GET /assays/ (paginated) → harvest all visible (id, title) pairs
          3. Intersect — keep only assays present in the project
          4. If a title appears multiple times, keep the lowest ID (oldest)
             and record duplicates in the cache's `_duplicates` block.
        """
        if verbose:
            print(f"  → GET /projects/{project_id}/ …", file=sys.stderr)
        t0 = time.monotonic()
        project = self.get_project(project_id)
        if verbose:
            print(f"    done in {time.monotonic()-t0:.2f}s", file=sys.stderr)
        rels = (project.get("data", {}).get("relationships", {})
                .get("assays", {}).get("data") or [])
        project_assay_ids = {str(item.get("id")) for item in rels if item.get("id")}
        if verbose:
            print(f"  → project has {len(project_assay_ids)} assay relationships",
                  file=sys.stderr)

        if not project_assay_ids:
            return {}

        if verbose:
            print(f"  → GET /assays/ (paginated) …", file=sys.stderr)
        # Collect all id→title from /assays/, then keep project-scoped ones
        by_title: Dict[str, list] = {}
        for aid, title in self.list_assays_paginated(verbose=verbose):
            if aid in project_assay_ids:
                by_title.setdefault(title, []).append(int(aid))

        result: Dict[str, int] = {}
        duplicates: Dict[str, list] = {}
        for title, ids in by_title.items():
            ids_sorted = sorted(ids)
            result[title] = ids_sorted[0]
            if len(ids_sorted) > 1:
                duplicates[title] = ids_sorted

        # Attach a metadata block for inspection if duplicates exist.
        # (Stripped before caller use — caller gets a clean title→id map.)
        if duplicates:
            result["__duplicates__"] = duplicates  # type: ignore[assignment]
        return result


# ─── CLI ─────────────────────────────────────────────────────────────────────

def _load_dotenv():
    """setdefault env vars from cwd/.env then <plugin>/.env (idempotent).

    Mirrors scripts/fdh/fdh_api.py:159-169, the reference implementation.
    Skips lines that are blank or start with '#'. Strips surrounding quotes.
    Existing env vars take precedence (we only setdefault).
    """
    for candidate in (Path.cwd() / ".env", _PLUGIN / ".env"):
        if not candidate.exists():
            continue
        for raw in candidate.read_text().splitlines():
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def cmd_fetch_assays(args: argparse.Namespace) -> int:
    _load_dotenv()

    username = args.username or os.environ.get("NEXTSEEK_USERNAME")
    password = args.password or os.environ.get("NEXTSEEK_PASSWORD")
    token = args.token or os.environ.get("NEXTSEEK_TOKEN")

    if not (username and password) and not token:
        print("error: provide credentials via one of:\n"
              "  --username + --password\n"
              "  NEXTSEEK_USERNAME + NEXTSEEK_PASSWORD env vars (or in .env)\n"
              "  --token / NEXTSEEK_TOKEN",
              file=sys.stderr)
        return 2

    client = NExtSEEKClient(username=username, password=password, token=token,
                            base_url=args.base_url)
    print(f"Fetching assays for project {args.project_id} from {args.base_url} "
          f"(auth: {client.auth_mode})…", file=sys.stderr)
    try:
        id_map = client.fetch_assay_id_map(args.project_id)
    except NExtSEEKError as e:
        print(f"\nAPI error: HTTP {e.status} on {e.url}\n{e.body[:1000]}",
              file=sys.stderr)
        return 1

    duplicates = id_map.pop("__duplicates__", None)

    if args.output:
        out_path = Path(args.output)
    else:
        try:
            cfg = config_from_args(args)
        except ProjectRootError as exc:
            print(f"error: {exc}", file=sys.stderr)
            return 2
        out_path = cfg.context / "assay_ids_cache.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "project_id": args.project_id,
        "base_url": args.base_url,
        "fetched_at_utc": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "assay_id_by_title": id_map,
    }
    if duplicates:
        payload["duplicate_titles"] = duplicates

    out_path.write_text(json.dumps(payload, indent=2, sort_keys=True))
    print(f"  ✓ wrote {len(id_map)} assays to {out_path}", file=sys.stderr)
    if duplicates:
        print(f"  ⚠ {len(duplicates)} titles have multiple IDs in this project "
              f"(see duplicate_titles in the cache file)", file=sys.stderr)
    return 0


def _client_from_args(args: argparse.Namespace) -> "NExtSEEKClient":
    """Resolve credentials the same way cmd_fetch_assays does, then build a client.

    Raises SystemExit(2) with the same guidance message when nothing is available.
    """
    _load_dotenv()
    username = getattr(args, "username", None) or os.environ.get("NEXTSEEK_USERNAME")
    password = getattr(args, "password", None) or os.environ.get("NEXTSEEK_PASSWORD")
    token = getattr(args, "token", None) or os.environ.get("NEXTSEEK_TOKEN")
    if not (username and password) and not token:
        print("error: provide credentials via one of:\n"
              "  --username + --password\n"
              "  NEXTSEEK_USERNAME + NEXTSEEK_PASSWORD env vars (or in .env)\n"
              "  --token / NEXTSEEK_TOKEN",
              file=sys.stderr)
        raise SystemExit(2)
    return NExtSEEKClient(username=username, password=password, token=token,
                          base_url=getattr(args, "base_url", DEFAULT_BASE_URL))


def _unwrap(doc: dict) -> dict:
    """JSON:API responses arrive as {'data': {...}} or {'data': [{...}]}."""
    data = doc.get("data", doc)
    return data[0] if isinstance(data, list) and data else data


def _attr_rows(record: dict) -> list:
    return (record.get("attributes") or {}).get("sample_attributes") or []


def _fmt_attr(a: dict) -> str:
    t = (a.get("sample_attribute_type") or {}).get("title") or "?"
    req = "required" if a.get("required") else "optional"
    return f"{str(a.get('title')):28s} {t:12s} {req:9s} (id={a.get('id')}, pos={a.get('pos')})"


def cmd_sampletype_get(args: argparse.Namespace) -> int:
    client = _client_from_args(args)
    rec = _unwrap(client.get_sample_type(args.sampletype))
    attrs = _attr_rows(rec)
    title = (rec.get("attributes") or {}).get("title")
    print(f"SampleType {title!r}  (id {rec.get('id')})  — {len(attrs)} attributes\n")
    for a in sorted(attrs, key=lambda x: x.get("pos") or 0):
        print(f"  {_fmt_attr(a)}")
    return 0


# Hosts where an --apply is a production write. The retired shim refused these
# without a second flag (`_confirm_production`), and dropping that guard when it
# was replaced left prose as the only thing between a curator and a global
# shared-schema write. A machine refusal is this plugin's house style for a
# production write -- `/curate-assay-write` sits behind eight of them.
PRODUCTION_HOSTS = ("nextseek.mit.edu",)


def _confirm_production(base_url: str, yes: bool) -> None:
    """Raise unless a production --apply carries --yes-production too."""
    host = str(base_url).split("//")[-1].split("/")[0].split(":")[0].lower()
    if host.rstrip(".") in PRODUCTION_HOSTS and not yes:
        raise SystemExit(
            f"REFUSED: {host} is production and --apply was given without "
            f"--yes-production.\n"
            f"  This is a GLOBAL, SHARED-SCHEMA write: it changes the type for "
            f"every project and every existing record of it.\n"
            f"  Rehearse on dev first:  --base-url https://nextseek-dev.mit.edu "
            f"(after the subcommand)")


def _report_automatic_changes(result: dict) -> None:
    """Print the side effects a request that 'adds one attribute' also made.

    MEASURED ON PRODUCTION: creating ONE attribute on BLD emitted 68
    `position_changed` automatic changes, renumbering every definition from
    position 8 down. Deleting the attribute again does NOT undo them. The count
    is in the response and is not obvious from the request, so it is surfaced
    rather than left for a reader to notice in the JSON.
    """
    changes = result.get("automatic_changes") or []
    if not isinstance(changes, list) or not changes:
        return
    kinds: dict[str, int] = {}
    for c in changes:
        kinds[str(c.get("type", "?")) if isinstance(c, dict) else "?"] = \
            kinds.get(str(c.get("type", "?")) if isinstance(c, dict) else "?", 0) + 1
    print(f"\n  AUTOMATIC CHANGES: {len(changes)} "
          f"({', '.join(f'{k}={v}' for k, v in sorted(kinds.items()))})",
          file=sys.stderr)
    print("  These are side effects of the request, not things you asked for, "
          "and a later delete does NOT undo them.", file=sys.stderr)


def cmd_attributes_list(args: argparse.Namespace) -> int:
    """GET /attributes/ — read the attributes the server actually holds.

    READS NEED ONLY A SEEK LOGIN, not a superuser. This is the authoritative
    answer to "what does this type accept"; the bundled catalog is a snapshot.
    """
    client = _client_from_args(args)
    try:
        result = client.list_attributes(args.sample_type)
    except NExtSEEKError as exc:
        print(f"FAILED ({exc.status}): {exc.body[:600]}", file=sys.stderr)
        return 1
    print(json.dumps(result, indent=2))
    return 0


def cmd_sampletype_remove_attribute(args: argparse.Namespace) -> int:
    """Remove an attribute from a live sample type.

    EXISTS BECAUSE THE REHEARSAL LOOP NEEDS IT. Step 3 of `/curate-sampletype
    apply` writes to dev for real; without a remove verb every rehearsal leaves
    a permanent attribute on dev's shared schema and the loop never closes.

    DESTRUCTIVE AND GLOBAL: it removes the field from every record of the type,
    and there is no undo through this API. Same production refusal as `add`.
    """
    client = _client_from_args(args)
    dry = not args.apply
    if not dry:
        _confirm_production(args.base_url, getattr(args, "yes_production", False))

    print(f"{'PLAN' if dry else 'APPLY'}: REMOVE {args.name!r} from sample type "
          f"{args.sampletype}", file=sys.stderr)
    if not dry:
        print("  This deletes the field from every existing record of this "
              "type. There is no undo.", file=sys.stderr)
    try:
        result = client.delete_attributes(args.sampletype,
                                          [{"title": args.name}], dry_run=dry)
    except NExtSEEKError as exc:
        print(f"REFUSED ({exc.status}): {exc.body[:600]}", file=sys.stderr)
        return 1
    print(json.dumps(result, indent=2))
    _report_automatic_changes(result)
    if dry:
        print("\nDRY RUN -- nothing removed. Re-run with --apply to commit.",
              file=sys.stderr)
    return 0


def cmd_sampletype_add_attribute(args: argparse.Namespace) -> int:
    """Add an attribute to a live sample type through the attributes API.

    THIS USED TO BE A DEAD END. `PATCH /nextseek_api/sample_types/{id}/` is a
    1:1 pass-through to SEEK, and SEEK enforces

        # lib/seek/samples/sample_type_editing_constraints.rb
        def allow_new_attribute?
          !samples?

    so it returns 422 for any type that already has samples, which NExtSEEK's
    proxy surfaces as a generic 502. That is still true of THAT route and is
    why this command no longer uses it. `POST /attributes/batch-create/` is a
    different, purpose-built route and is not subject to the constraint.

    DRY RUN BY DEFAULT. `--apply` is required to write, and the preview is the
    server's own plan rather than this client's guess at one.
    """
    client = _client_from_args(args)
    attribute = {"title": args.name, "sample_attribute_type": args.type}
    if args.required:
        attribute["required"] = True

    dry = not args.apply
    if not dry:
        _confirm_production(args.base_url, getattr(args, "yes_production", False))
    if args.debug:
        print(json.dumps(client._targets(args.sampletype, [attribute])
                         | {"dry_run": dry}, indent=2), file=sys.stderr)

    print(f"{'PLAN' if dry else 'APPLY'}: add {args.name!r} ({args.type}) "
          f"to sample type {args.sampletype}", file=sys.stderr)
    if not dry:
        print("  GLOBAL, SHARED-SCHEMA WRITE: this changes the type for every "
              "project and every existing record of it.", file=sys.stderr)

    try:
        result = client.create_attributes(args.sampletype, [attribute],
                                          dry_run=dry)
    except NExtSEEKError as exc:
        # The API answers in structured JSON; surface it rather than a status.
        print(f"REFUSED ({exc.status}): {exc.body[:600]}", file=sys.stderr)
        if exc.status == 403:
            print("  Mutations need a Django SUPERUSER. A SEEK admin who is "
                  "not one is refused here -- the two populations are not "
                  "nested.", file=sys.stderr)
        return 1

    print(json.dumps(result, indent=2))
    _report_automatic_changes(result)
    if dry:
        print("\nDRY RUN -- nothing written. Re-run with --apply to commit.",
              file=sys.stderr)
        print("  Read AUTOMATIC CHANGES above before you do.", file=sys.stderr)
    else:
        print("\nWritten. No worker restart is required: the attribute cache "
              "is invalidated by a database-side generation stamp read once "
              "per batch.", file=sys.stderr)
    return 0


def _cmd_sampletype_add_attribute_dead(args: argparse.Namespace) -> int:
    """Original REST implementation. Unreachable; kept for reference only."""
    client = _client_from_args(args)
    rec = _unwrap(client.get_sample_type(args.sampletype))
    type_id = rec.get("id")
    existing = _attr_rows(rec)
    names = {str(a.get("title")) for a in existing}

    print(f"SampleType {args.sampletype!r} (id {type_id}) — {len(existing)} attributes today")

    if args.name in names:
        print(f"\n✓ {args.name!r} is ALREADY defined — nothing to do.")
        return 0

    # Rebuild the complete array. Existing entries keep their id so the server
    # updates them in place instead of recreating (and renumbering) them.
    #
    # sample_attribute_type is sent as an ID ref, not a title ref: the GET returns
    # a rich object ({id, title, base_type, regexp}) and the upstream SEEK is
    # happier resolving the id it already issued than re-matching on title.
    # `description` and `pid` are echoed back so a PATCH does not blank them.
    def _type_ref(a: dict) -> dict:
        t = a.get("sample_attribute_type") or {}
        return {"id": str(t["id"])} if t.get("id") else {"title": t.get("title") or "Text"}

    payload = []
    for a in sorted(existing, key=lambda x: x.get("pos") or 0):
        row = {
            "id": str(a.get("id")),
            "title": a.get("title"),
            "sample_attribute_type": _type_ref(a),
            "required": bool(a.get("required")),
            "pos": a.get("pos"),
        }
        if a.get("description") is not None:
            row["description"] = a["description"]
        if a.get("pid") is not None:
            row["pid"] = a["pid"]
        payload.append(row)

    # Resolve the new attribute's type id from an existing attribute of the same
    # title, so we reuse the server's own id rather than relying on title matching.
    type_ref = {"title": args.type}
    for a in existing:
        t = a.get("sample_attribute_type") or {}
        if t.get("title") == args.type and t.get("id"):
            type_ref = {"id": str(t["id"])}
            break

    next_pos = max([a.get("pos") or 0 for a in existing] or [0]) + 1
    payload.append({                     # no id -> server creates it
        "title": args.name,
        "sample_attribute_type": type_ref,
        "required": bool(args.required),
        "pos": next_pos,
    })

    print(f"\nPLAN: add {args.name!r} ({args.type}, "
          f"{'required' if args.required else 'optional'}) at pos {next_pos}")
    print(f"      re-sending {len(existing)} existing attributes unchanged "
          f"(server replaces the array, so they must all be present)")

    if getattr(args, "debug", False):
        import json as _j
        print("\n--- payload that would be sent ---")
        print(_j.dumps({"data": {"id": str(type_id), "type": "sample_types",
                                 "attributes": {"sample_attributes": payload}}}, indent=1)[:2500])

    if not args.apply:
        print("\nDRY RUN — nothing sent. Re-run with --apply to PATCH.")
        print("NOTE: sample types are GLOBAL. This affects every project and every")
        print("      existing record of this type across NExtSEEK.")
        return 0

    print(f"\nPATCHing {args.sampletype} …")
    client.patch_sample_type(type_id, payload)

    # Trust the round-trip, not the response body.
    after = _attr_rows(_unwrap(client.get_sample_type(args.sampletype)))
    after_names = {str(a.get("title")) for a in after}
    lost = names - after_names
    if lost:
        print(f"  ✗ ATTRIBUTES LOST: {sorted(lost)} — investigate immediately")
        return 1
    if args.name not in after_names:
        print(f"  ✗ {args.name!r} not present after PATCH")
        return 1
    print(f"  ✓ verified: {len(after)} attributes, {args.name!r} added, "
          f"all {len(existing)} originals intact")
    return 0


def cmd_validate(args: argparse.Namespace) -> int:
    """Dry-run validate one or more xlsx files against NExtSEEK without inserting."""
    _load_dotenv()

    username = args.username or os.environ.get("NEXTSEEK_USERNAME")
    password = args.password or os.environ.get("NEXTSEEK_PASSWORD")
    token = args.token or os.environ.get("NEXTSEEK_TOKEN")

    if not (username and password) and not token:
        print("error: provide credentials via --username/--password, .env, "
              "or --token / NEXTSEEK_TOKEN", file=sys.stderr)
        return 2

    client = NExtSEEKClient(username=username, password=password, token=token,
                            base_url=args.base_url, timeout=120.0)

    files = [Path(p) for p in args.files]
    for fp in files:
        if not fp.is_file():
            print(f"error: file not found: {fp}", file=sys.stderr)
            return 2

    print(f"Validating {len(files)} file(s) against {args.base_url} "
          f"(project {args.project_id}, checks={args.checks}, "
          f"auth={client.auth_mode})\n", file=sys.stderr)

    overall_valid = True
    for fp in files:
        print(f"━━ {fp.name} ━━")
        try:
            result = client.validate_batch_upload(
                fp, project_id=args.project_id, checks=args.checks)
        except NExtSEEKError as e:
            print(f"  ✗ API error: HTTP {e.status}\n    {e.body[:500]}")
            overall_valid = False
            continue

        valid = result.get("valid", False)
        summary = result.get("summary", "")
        totals = result.get("totals") or {}
        errors = result.get("errors") or []
        warnings = result.get("warnings") or {}
        checks_run = result.get("checks_run") or []
        checks_skipped = result.get("checks_skipped") or []

        flag = "✓ VALID" if valid else "✗ INVALID"
        print(f"  {flag}  — {summary}")
        if totals:
            print(f"  totals: processed={totals.get('processed', '?')}, "
                  f"success={totals.get('success', '?')}, "
                  f"failed={totals.get('failed', '?')}, "
                  f"skipped={totals.get('skipped', '?')}")
        if checks_run:
            print(f"  checks_run: {', '.join(checks_run)}")
        if checks_skipped:
            print(f"  checks_skipped: {', '.join(checks_skipped)}")
        if errors:
            print(f"  ERRORS ({len(errors)}):")
            for e in errors[:20]:
                etype = e.get("type", "?")
                emsg = e.get("message", "?")
                print(f"    [{etype}] {emsg}")
            if len(errors) > 20:
                print(f"    ... and {len(errors) - 20} more")
        if warnings:
            n = len(warnings) if isinstance(warnings, dict) else 0
            print(f"  warnings: {n} group(s)")
            for k, v in list(warnings.items())[:10]:
                vs = v if isinstance(v, str) else json.dumps(v)[:200]
                print(f"    {k}: {vs}")

        # Optional: dump full result to JSON for inspection
        if args.dump_dir:
            dump_path = Path(args.dump_dir) / f"{fp.stem}.validate.json"
            dump_path.parent.mkdir(parents=True, exist_ok=True)
            dump_path.write_text(json.dumps(result, indent=2, sort_keys=True))
            print(f"  → full response: {dump_path}")

        if not valid:
            overall_valid = False
        print()

    print(f"{'━'*60}")
    print(f"{'ALL FILES VALID ✓' if overall_valid else 'SOME FILES INVALID ✗'}")
    return 0 if overall_valid else 1


def cmd_pull_db(args: argparse.Namespace) -> int:
    """Download a project's full DB export into previous_metadata/.

    GET /admin/project-export/{id}/?output_format=xlsx → the master workbook.
    This is the fresh DB pull the build stamp-guard requires before minting.
    """
    _load_dotenv()
    client = _client_from_args(args)  # exits(2) if no creds

    if args.dest:
        dest_dir = Path(args.dest)
    else:
        try:
            dest_dir = config_from_args(args).root / "previous_metadata"
        except ProjectRootError as exc:
            print(f"error: {exc}\n  Run inside a curation project, or pass --dest.",
                  file=sys.stderr)
            return 2
    dest_dir.mkdir(parents=True, exist_ok=True)

    print(f"Pulling project {args.project_id} export ({args.output_format}) "
          f"from {args.base_url} …")
    try:
        content, server_name = client.export_project(args.project_id,
                                                      args.output_format)
    except NExtSEEKError as e:
        print(f"\nAPI error: HTTP {e.status} on {e.url}\n{e.body[:800]}",
              file=sys.stderr)
        return 1

    out = dest_dir / (args.filename or server_name)
    out.write_bytes(content)
    print(f"  ✓ wrote {out}  ({len(content):,} bytes)")

    if args.output_format == "xlsx":
        try:
            import io
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            total = sum(max(sum(1 for _ in wb[sh].iter_rows()) - 1, 0)
                        for sh in wb.sheetnames)
            print(f"  {len(wb.sheetnames)} sheets, ~{total} data rows")
        except Exception as e:  # introspection is best-effort
            print(f"  (workbook introspection skipped: {e})")
    print("  → the build stamp-guard reads the newest xlsx here as the fresh "
          "pull.")
    return 0


def cmd_detect_context(args: argparse.Namespace) -> int:
    """Suggest project + lab + pi for /curate-init as JSON (API + local evidence)."""
    import json as _json
    _load_dotenv()
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import detect_context as dc

    client = _client_from_args(args)  # exits(2) if no creds
    try:
        cfg = config_from_args(args)
        root, prev_dir = cfg.root, cfg.previous_metadata
    except ProjectRootError:
        root = Path(".").resolve()
        prev_dir = root / "previous_metadata"

    warnings = []
    evidence = dc.gather_evidence(root)
    try:
        projects = client.list_projects()
    except NExtSEEKError as e:
        print(f"error: could not list projects: HTTP {e.status}", file=sys.stderr)
        return 1
    ranked = dc.rank_projects(projects, evidence)

    if getattr(args, "project_id", None):
        chosen = next((p for p in ranked if str(p["id"]) == str(args.project_id)),
                      {"id": args.project_id, "title": ""})
    else:
        chosen = ranked[0] if ranked else None

    labs, export_path = [], None
    if chosen:
        try:
            content, fname = client.export_project(chosen["id"], "xlsx")
            prev_dir.mkdir(parents=True, exist_ok=True)
            (prev_dir / fname).write_bytes(content)
            export_path = str(prev_dir / fname)
            labs = dc.rank_labs(dc.extract_labs(content), evidence)
        except NExtSEEKError as e:
            warnings.append(f"export pull failed (HTTP {e.status}); labs unavailable")

    out = {
        "projects": ranked,
        "chosen_project": chosen,
        "labs": [vars(l) for l in labs],
        "pi_guess": dc.guess_pi(labs, evidence, getattr(args, "pi", None)),
        "export_path": export_path,
        "evidence": {
            "path_tokens": evidence.path_tokens[:20],
            "author_surnames": evidence.author_surnames[:20],
            "master_tokens": evidence.master_tokens[:20],
        },
        "warnings": warnings,
    }
    print(_json.dumps(out, indent=2))
    return 0


def main(argv=None) -> int:
    p = argparse.ArgumentParser(
        description="NExtSEEK API helper for resolving assay titles → IDs.")
    sub = p.add_subparsers(dest="cmd", required=True)

    # ── fetch-assays ────────────────────────────────────────────────────────
    fa = sub.add_parser(
        "fetch-assays",
        help="Fetch project-scoped assay title→id map and cache it locally.")
    fa.add_argument("--project-id", required=True,
                    help="SEEK project ID (numeric) or NExtSEEK UID (string).")
    fa.add_argument("--username", default=None,
                    help="Basic auth username. If omitted, reads $NEXTSEEK_USERNAME "
                         "(also auto-loaded from REPO/.env).")
    fa.add_argument("--password", default=None,
                    help="Basic auth password. If omitted, reads $NEXTSEEK_PASSWORD. "
                         "CLI use is discouraged (shell history); prefer .env.")
    fa.add_argument("--token", default=None,
                    help="API token (alternative to basic auth). "
                         "If omitted, reads $NEXTSEEK_TOKEN.")
    fa.add_argument("--base-url", default=DEFAULT_BASE_URL,
                    help=f"NExtSEEK API base URL (default: {DEFAULT_BASE_URL}).")
    fa.add_argument("--output", type=Path, default=None,
                    help="Where to write assay_ids_cache.json "
                         "(default: <project-root>/context/assay_ids_cache.json)")
    add_config_args(fa)
    fa.set_defaults(func=cmd_fetch_assays)

    # ── pull-db ─────────────────────────────────────────────────────────────
    pd = sub.add_parser(
        "pull-db",
        help="Download a project's full DB export (xlsx) into "
             "previous_metadata/ — the fresh pull the build stamp-guard checks.")
    pd.add_argument("--project-id", required=True,
                    help="SEEK project ID (numeric), e.g. 10 for CSBC.")
    pd.add_argument("--output-format", default="xlsx", choices=["xlsx", "json"],
                    help="Export format (default: xlsx — the master workbook).")
    pd.add_argument("--dest", default=None,
                    help="Destination directory "
                         "(default: <project-root>/previous_metadata/).")
    pd.add_argument("--filename", default=None,
                    help="Override output filename (default: server-provided name).")
    pd.add_argument("--username", default=None)
    pd.add_argument("--password", default=None)
    pd.add_argument("--token", default=None)
    pd.add_argument("--base-url", default=DEFAULT_BASE_URL,
                    help=f"NExtSEEK API base URL (default: {DEFAULT_BASE_URL}).")
    add_config_args(pd)
    pd.set_defaults(func=cmd_pull_db)

    # ── detect-context ──────────────────────────────────────────────────────
    dctx = sub.add_parser(
        "detect-context",
        help="Suggest project + lab code + pi for /curate-init (JSON) from the "
             "API and local evidence.")
    dctx.add_argument("--project-id", default=None,
                      help="Force a project id instead of auto-ranking.")
    # NOTE: --pi is NOT added here — add_config_args() below already registers
    # it ("PI short name") for the project-config group, and cmd_detect_context
    # reads it the same way (getattr(args, "pi", None): "Known PI, skips
    # guessing"). Adding it twice raises argparse.ArgumentError (conflicting
    # option string) at parser build time.
    dctx.add_argument("--username", default=None)
    dctx.add_argument("--password", default=None)
    dctx.add_argument("--token", default=None)
    dctx.add_argument("--base-url", default=DEFAULT_BASE_URL)
    add_config_args(dctx)
    dctx.set_defaults(func=cmd_detect_context)

    # ── validate ────────────────────────────────────────────────────────────
    va = sub.add_parser(
        "validate",
        help="Dry-run validate xlsx files against NExtSEEK (no INSERT, no side effects).")
    va.add_argument("--project-id", required=True,
                    help="SEEK project ID — required even for dry-run.")
    va.add_argument("files", nargs="+",
                    help="One or more .xlsx files to validate.")
    va.add_argument("--checks", default="structure",
                    help="Comma-separated subset of: structure, name_check, dag. "
                         "Default: structure (fastest). Use 'structure,dag,name_check' "
                         "for the most thorough check.")
    va.add_argument("--dump-dir", default=None,
                    help="If set, write each file's full ValidationResult JSON here.")
    va.add_argument("--username", default=None)
    va.add_argument("--password", default=None)
    va.add_argument("--token", default=None)
    va.add_argument("--base-url", default=DEFAULT_BASE_URL)
    add_config_args(va)
    va.set_defaults(func=cmd_validate)

    # ── sample-type schema inspection / patching ────────────────────────────
    stg = sub.add_parser(
        "sampletype-get",
        help="Show a sample type's current attribute list (read-only).")
    stg.add_argument("sampletype", help="Short code or numeric id, e.g. A.TITR")
    stg.add_argument("--username", default=None)
    stg.add_argument("--password", default=None)
    stg.add_argument("--token", default=None)
    stg.add_argument("--base-url", default=DEFAULT_BASE_URL)
    stg.set_defaults(func=cmd_sampletype_get)

    al = sub.add_parser(
        "attributes-list",
        help="Read sample-type attributes from the server (read-only)")
    al.add_argument("--sample-type", default=None,
                    help="Restrict to one type: id, numeric string, or exact title")
    al.add_argument("--username", default=None)
    al.add_argument("--password", default=None)
    al.add_argument("--token", default=None)
    al.add_argument("--base-url", default=DEFAULT_BASE_URL)
    al.set_defaults(func=cmd_attributes_list)

    sta = sub.add_parser(
        "sampletype-add-attribute",
        help="Add an attribute to a live sample type (attributes API; "
             "dry run unless --apply)")
    sta.add_argument("sampletype", help="Short code or numeric id, e.g. A.TITR")
    sta.add_argument("--name", required=True, help="Attribute title, e.g. Notes")
    sta.add_argument("--type", default="Text",
                     help="SampleAttributeType title (default: Text)")
    sta.add_argument("--required", action="store_true",
                     help="Mark the new attribute required (default: optional)")
    sta.add_argument("--debug", action="store_true",
                     help="Print the exact JSON payload before sending.")
    sta.add_argument("--apply", action="store_true",
                     help="Actually write. Without this the server plans the "
                          "change and writes nothing (dry_run).")
    sta.add_argument("--yes-production", action="store_true",
                     help="Required IN ADDITION to --apply when the target is "
                          "production. Rehearse on dev first.")
    sta.add_argument("--username", default=None)
    sta.add_argument("--password", default=None)
    sta.add_argument("--token", default=None)
    sta.add_argument("--base-url", default=DEFAULT_BASE_URL)
    sta.set_defaults(func=cmd_sampletype_add_attribute)

    strm = sub.add_parser(
        "sampletype-remove-attribute",
        help="Remove an attribute from a live sample type "
             "(DESTRUCTIVE; dry run unless --apply)")
    strm.add_argument("sampletype", help="Short code or numeric id, e.g. A.TITR")
    strm.add_argument("--name", required=True, help="Attribute title to remove")
    strm.add_argument("--apply", action="store_true",
                      help="Actually remove. Without this the server plans it "
                           "and removes nothing (dry_run).")
    strm.add_argument("--yes-production", action="store_true",
                      help="Required IN ADDITION to --apply when the target is "
                           "production.")
    strm.add_argument("--username", default=None)
    strm.add_argument("--password", default=None)
    strm.add_argument("--token", default=None)
    strm.add_argument("--base-url", default=DEFAULT_BASE_URL)
    strm.set_defaults(func=cmd_sampletype_remove_attribute)

    args = p.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
