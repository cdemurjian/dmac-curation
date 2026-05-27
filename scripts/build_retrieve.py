# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Build RETRIEVE.TXT from assay_sheets/ — newline-separated UIDs for chat_nextseek.

By default emits only downstream sample types (D.*/A.*/SLD/etc.) — the retrieve
function auto-pulls parents via the lineage chain. Use --include-parents to emit
all UIDs including DNA/RNA/TIS/MUS intermediates.

Prefers `*-upload-new.xlsx` over `*-upload.xlsx` to capture the latest curation.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

# Sample types treated as PARENTS (auto-pulled by retrieve; excluded by default)
PARENT_TYPES = {"MUS", "TIS", "DNA", "RNA", "PAT", "PAV", "CHM", "CEL"}


def collect_uids(assay_sheets_dir: Path, include_parents: bool) -> list[str]:
    """Walk assay_sheets/, prefer -upload-new over -upload, dedupe + sort UIDs."""
    seen: set[str] = set()

    # Build map of basename → preferred file (prefer -upload-new)
    candidates: dict[str, Path] = {}
    for p in sorted(assay_sheets_dir.glob("*.xlsx")):
        if p.name.startswith("~"):  # openpyxl lock files
            continue
        if "-upload-new" in p.stem:
            base = p.stem.replace("-upload-new", "")
            candidates[base] = p
        elif "-upload" in p.stem and p.stem.replace("-upload", "") not in candidates:
            base = p.stem.replace("-upload", "")
            candidates.setdefault(base, p)

    for path in candidates.values():
        wb = load_workbook(path, read_only=True, data_only=True)
        if "Samples" not in wb.sheetnames:
            continue
        ws = wb["Samples"]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            continue
        uid_col = None
        for i, h in enumerate(header):
            if h and str(h).strip().lower() == "uid":
                uid_col = i
                break
        if uid_col is None:
            continue
        for row in rows_iter:
            if uid_col >= len(row):
                continue
            uid = row[uid_col]
            if not uid:
                continue
            uid = str(uid).strip()
            if "-" not in uid:
                continue
            sample_type = uid.split("-", 1)[0]
            if not include_parents and sample_type in PARENT_TYPES:
                continue
            seen.add(uid)

    return sorted(seen)


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--assay-sheets", default="assay_sheets",
                   help="Directory containing *-upload-new.xlsx / *-upload.xlsx (default: assay_sheets)")
    p.add_argument("--output", default="RETRIEVE.TXT", help="Output file (default: RETRIEVE.TXT)")
    p.add_argument("--include-parents", action="store_true",
                   help="Include MUS/TIS/DNA/RNA/PAT/PAV/CHM/CEL UIDs (default: downstream-only)")
    args = p.parse_args()

    assay_dir = Path(args.assay_sheets).resolve()
    if not assay_dir.is_dir():
        print(f"ERROR: {assay_dir} is not a directory", file=sys.stderr)
        return 2

    uids = collect_uids(assay_dir, include_parents=args.include_parents)
    out = Path(args.output).resolve()
    out.write_text("\n".join(uids) + "\n")

    by_type: dict[str, int] = {}
    for u in uids:
        t = u.split("-", 1)[0]
        by_type[t] = by_type.get(t, 0) + 1

    print(f"Wrote {len(uids)} UIDs to {out}")
    for t in sorted(by_type):
        print(f"  {t}: {by_type[t]}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
