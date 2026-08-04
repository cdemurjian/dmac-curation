# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Pre-mint collision guard for /curate-build.

Root-cause fix for the UID-stamp collision class of bug: a build script mints
UIDs from N=1 under a ``<YYMMDD><LAB>`` stamp that another curation batch already
claimed, and the upload silently OVERWRITES the other study's records.

Two gates a build script MUST pass before minting (call ``preflight``):

  1. require_fresh_db_pull(...)
     There must be a DB pull (a master/AllMetadata export) that is present AND
     recent — pulled ~this session. A stale pull is worse than none: it will not
     show a stamp another curator claimed since it was exported, so the stamp
     check would pass falsely. Missing or stale -> StampGuardError.

  2. guard_stamp(...)
     Collect every UID in the DB pull under the intended ``<date><lab>`` stamp,
     per sample type. If ANY sample type the batch will mint already has rows
     under that stamp, refuse — with the used ranges and a suggested free stamp.
     Clean -> return.

Refuse-by-default (not auto-continue-after-max-N): interleaving two studies in
one stamp is exactly what produces fragile, half-overwritten batches, and two
scripts minting into the same stamp concurrently would race. One clean stamp per
batch is the invariant.

Usage in a build_<arm>.py (after the sys.path insert that already imports _common):

    from stamp_guard import preflight
    preflight(["CEL", "D.MSP", "A.MSP"], LAB, DATE, project_root=".")

Set ``STAMP_GUARD_OVERRIDE=1`` in the environment to downgrade the collision
error to a printed warning — for the deliberate, eyes-open case where you are
re-uploading into an existing stamp on purpose (e.g. a correction re-run).
"""
from __future__ import annotations

import datetime as _dt
import os
import re
from pathlib import Path

from openpyxl import load_workbook

UID_RE = re.compile(r"^(?P<type>.+)-(?P<date>\d{6})(?P<lab>[A-Z]{3})-(?P<n>\d+)$")


class StampGuardError(RuntimeError):
    """Raised when the pre-mint guard refuses to let a build proceed."""


# ── DB-pull discovery + freshness ──────────────────────────────────────────
def find_db_pull(project_root: str | Path = ".") -> Path | None:
    """Newest .xlsx in ``<project_root>/previous_metadata/`` (the DB-pull drop),
    or None if the directory is empty/absent. Ignores Excel lock files (~$…)."""
    pm = Path(project_root) / "previous_metadata"
    if not pm.is_dir():
        return None
    cands = [p for p in pm.glob("*.xlsx") if not p.name.startswith("~$")]
    if not cands:
        return None
    return max(cands, key=lambda p: p.stat().st_mtime)


def require_fresh_db_pull(
    master_path: str | Path | None = None,
    project_root: str | Path = ".",
    max_age_hours: float = 24.0,
) -> Path:
    """Ensure a DB pull exists and was exported recently. Returns its path.

    ``master_path`` — explicit file; if None, auto-discovers the newest xlsx in
    ``previous_metadata/``. Raises StampGuardError if none is found or the newest
    one is older than ``max_age_hours`` (a stale pull can't prove a stamp is free).
    """
    path = Path(master_path) if master_path else find_db_pull(project_root)
    if path is None or not Path(path).exists():
        raise StampGuardError(
            "No DB pull found. Drop a fresh NExtSEEK export (the 'CSBC All …' / "
            "AllMetadata xlsx) into previous_metadata/ before building — the stamp "
            "check needs current database state to be meaningful."
        )
    age_h = (_dt.datetime.now().timestamp() - Path(path).stat().st_mtime) / 3600.0
    if age_h > max_age_hours:
        raise StampGuardError(
            f"DB pull '{Path(path).name}' is {age_h:.1f} h old (> {max_age_hours} h). "
            "Pull a fresh export right before building; a stale pull won't show "
            "stamps another curator claimed since it was exported."
        )
    return Path(path)


# ── stamp-usage scan ───────────────────────────────────────────────────────
def scan_used(master_path: str | Path) -> dict[tuple[str, str], list[int]]:
    """Map ``(sampletype, '<date><lab>') -> [N,…]`` across every sheet of the pull."""
    used: dict[tuple[str, str], list[int]] = {}
    wb = load_workbook(master_path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        it = ws.iter_rows(values_only=True)
        try:
            hdr = [str(h).strip().lower() if h is not None else "" for h in next(it)]
        except StopIteration:
            continue
        if "uid" not in hdr:
            continue
        ui = hdr.index("uid")
        for r in it:
            if ui < len(r) and r[ui]:
                m = UID_RE.match(str(r[ui]).strip())
                if m:
                    key = (m["type"], f'{m["date"]}{m["lab"]}')
                    used.setdefault(key, []).append(int(m["n"]))
    wb.close()
    return used


def _suggest_free_stamp(used: dict, lab: str, date: str) -> str:
    """Suggest the nearest YYMMDD (same lab) with no UID of ANY type in the pull."""
    taken = {stamp[:6] for (_t, stamp) in used if stamp.endswith(lab)}
    try:
        base = _dt.datetime.strptime(date, "%y%m%d").date()
    except ValueError:
        base = _dt.date.today()
    for delta in range(0, 30):          # walk forward from the intended date
        cand = (base + _dt.timedelta(days=delta)).strftime("%y%m%d")
        if cand not in taken:
            return f"{cand}{lab}"
    return f"{date}{lab}-CHOOSE-A-FREE-STAMP"


def guard_stamp(
    master_path: str | Path,
    sample_types: list[str],
    lab: str,
    date: str,
) -> None:
    """Refuse if any ``sample_types`` already have UIDs under ``<date><lab>``.

    Raises StampGuardError (unless STAMP_GUARD_OVERRIDE=1) naming the collided
    types, their used N-ranges, and a suggested free stamp.
    """
    used = scan_used(master_path)
    stamp = f"{date}{lab}"
    hits = []
    for st in sample_types:
        ns = sorted(used.get((st, stamp), []))
        if ns:
            hits.append(f"{st}: {len(ns)} UIDs, N {ns[0]}..{ns[-1]}")
    if not hits:
        return
    free = _suggest_free_stamp(used, lab, date)
    msg = (
        f"STAMP COLLISION — '{stamp}' is already in use in {Path(master_path).name} "
        f"for sample type(s) you are about to mint:\n    "
        + "\n    ".join(hits)
        + f"\nMinting from N=1 would OVERWRITE those records on upload. "
        f"Re-mint this batch under a free stamp — suggested: {free} "
        f"(or any unused <YYMMDD>{lab})."
    )
    if os.environ.get("STAMP_GUARD_OVERRIDE") == "1":
        print(f"[stamp_guard] WARNING (override on): {msg}")
        return
    raise StampGuardError(msg)


def preflight(
    sample_types: list[str],
    lab: str,
    date: str,
    *,
    project_root: str | Path = ".",
    master_path: str | Path | None = None,
    max_age_hours: float = 24.0,
) -> Path:
    """Run both gates. Returns the DB-pull path used. Raises StampGuardError on
    a missing/stale pull or a stamp collision. Call once at the top of a build."""
    pull = require_fresh_db_pull(master_path, project_root, max_age_hours)
    guard_stamp(pull, sample_types, lab, date)
    print(f"[stamp_guard] OK — {date}{lab} is free for {sample_types} "
          f"(checked against {pull.name})")
    return pull
