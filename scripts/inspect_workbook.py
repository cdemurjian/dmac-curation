# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Inspect an .xlsx workbook — sheet names, dimensions, headers, sample rows.

Common idiom across past sessions; canonicalized here.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook


def inspect(path: Path, sheet: str | None, sample_rows: int) -> int:
    if not path.exists():
        print(f"ERROR: {path} not found", file=sys.stderr)
        return 2
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"ERROR: cannot open {path}: {e}", file=sys.stderr)
        return 2
    try:
        print(f"## {path.name}")
        print(f"Sheets ({len(wb.sheetnames)}):")
        for s in wb.sheetnames:
            ws = wb[s]
            # max_row/max_col are unreliable in read_only — count manually
            all_rows = list(ws.iter_rows(values_only=True))
            n_rows = len(all_rows)
            n_cols = max((len(r) for r in all_rows), default=0)
            print(f"  - {s}: {n_rows} rows (incl. header) × {n_cols} cols")

        missing = False
        targets = [sheet] if sheet else wb.sheetnames
        for s in targets:
            if s not in wb.sheetnames:
                print(f"\n(sheet '{s}' not found)")
                missing = True
                continue
            ws = wb[s]
            print(f"\n### Sheet: {s}")
            iter_rows = ws.iter_rows(values_only=True)
            header = next(iter_rows, None)
            if header is None:
                print("  (empty)")
                continue
            print(f"  Headers ({len(header)}): {list(header)}")
            if sample_rows > 0:
                print(f"  Sample (up to {sample_rows} rows):")
                for i, row in enumerate(iter_rows):
                    if i >= sample_rows:
                        break
                    print(f"    [{i}] {list(row)}")
        return 1 if missing else 0
    finally:
        wb.close()


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("path", type=Path, help="Path to .xlsx file")
    p.add_argument("--sheet", help="Limit detail to this sheet (default: all)")
    p.add_argument("--sample", type=int, default=0, help="Show this many sample data rows per sheet (0 = none)")
    args = p.parse_args()
    return inspect(args.path, args.sheet, args.sample)


if __name__ == "__main__":
    sys.exit(main())
