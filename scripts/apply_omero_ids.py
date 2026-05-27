#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Apply OMERO image IDs to D.IMG `Link_PrimaryData` columns.

Reads `omero_images.csv` (output of `omero_pull.py all --project N`) and patches
the matching upload sheet rows by filename. Idempotent.
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

from openpyxl import load_workbook


def apply(xlsx_path: Path, omero_csv: Path, dry_run: bool) -> int:
    by_filename: dict[str, dict[str, str]] = {}
    with omero_csv.open() as f:
        for row in csv.DictReader(f):
            fname = row.get("filename") or ""
            if fname:
                by_filename[fname] = row

    wb = load_workbook(xlsx_path)
    try:
        if "Samples" not in wb.sheetnames:
            print(f"ERROR: no Samples sheet in {xlsx_path}", file=sys.stderr)
            return 2
        ws = wb["Samples"]

        header_row = [c.value for c in ws[1]]
        try:
            file_col = header_row.index("File_PrimaryData") + 1
            link_col = header_row.index("Link_PrimaryData") + 1
        except ValueError:
            print("ERROR: File_PrimaryData and Link_PrimaryData columns required", file=sys.stderr)
            return 2

        patched = 0
        for r in range(2, ws.max_row + 1):
            fname_cell = ws.cell(row=r, column=file_col).value
            if not fname_cell:
                continue
            fname = str(fname_cell).strip()
            match = by_filename.get(fname)
            if not match:
                continue
            link = match.get("web_url") or match.get("show_url") or ""
            if not link:
                continue
            if not dry_run:
                ws.cell(row=r, column=link_col).value = link
            patched += 1
            print(f"  {'(dry-run) ' if dry_run else ''}{fname} → {link}")

        if not dry_run and patched > 0:
            wb.save(xlsx_path)
        print(f"\n{'Would patch' if dry_run else 'Patched'} {patched} rows in {xlsx_path}")
        return 0
    finally:
        wb.close()


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("xlsx", type=Path)
    p.add_argument("--omero-csv", type=Path, default=Path("omero_images.csv"))
    p.add_argument("--write", action="store_true", help="Apply changes (default is dry-run)")
    args = p.parse_args()
    return apply(args.xlsx, args.omero_csv, dry_run=not args.write)


if __name__ == "__main__":
    sys.exit(main())
