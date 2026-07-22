#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
# Lifted from lee/scripts/apply_zenodo_links.py; generalized for dmac-curation plugin.
"""Apply Zenodo zip URLs to the matching rows in assay_sheets/*-upload*.xlsx.

For each zip in Zenodo_upload/ (or --zip-dir):
  - Walk its contents
  - Look up each contained filename in the metadata workbook → (sample_type, UID)
  - Write the zip's URL into Link_PrimaryData for that UID's row in the right upload sheet

Usage:
  uv run scripts/apply_zenodo_links.py --record-id 12345678            # preview (default)
  uv run scripts/apply_zenodo_links.py --record-id 12345678 --write    # patch the sheets
"""
from __future__ import annotations

import argparse
import shutil
import sys
import zipfile
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _config import add_config_args, config_from_args  # noqa: E402
from _config import ProjectRootError  # noqa: E402

# TODO(v0.2): support loading SHEET_FOR_TYPE from a project-level config file.
# The default below is empty; callers must supply assay sheet paths via --sheet-map
# or rely on the auto-discovery glob (any *-upload*.xlsx in assay_sheets/).
SHEET_FOR_TYPE: dict[str, Path] = {}


def load_curation_index(meta: Path) -> dict[str, tuple[str, str]]:
    """Build {filename → (sample_type, UID)} from all sheets in the metadata workbook."""
    wb = load_workbook(meta, read_only=True, data_only=True)
    try:
        out: dict[str, tuple[str, str]] = {}
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            hdr = None
            for r in ws.iter_rows(values_only=True):
                if hdr is None:
                    hdr = list(r)
                    i_uid = hdr.index("UID") if "UID" in hdr else None
                    i_file = hdr.index("File_PrimaryData") if "File_PrimaryData" in hdr else None
                    if i_uid is None or i_file is None:
                        break
                    continue
                if r[i_file]:
                    fn = str(r[i_file]).strip()
                    if fn not in out:
                        out[fn] = (sheet, str(r[i_uid]))
        return out
    finally:
        wb.close()


def discover_sheet_map(assay_dir: Path) -> dict[str, Path]:
    """Map sample type → sheet path. Prefer -upload-new over -upload."""
    by_base: dict[str, Path] = {}
    for p in sorted(assay_dir.glob("*.xlsx")):
        if "-upload-new" in p.stem:
            base = p.stem.replace("-upload-new", "")
            by_base[base] = p  # always overrides
        elif "-upload" in p.stem:
            base = p.stem.replace("-upload", "")
            by_base.setdefault(base, p)  # only set if not already from -new
    return by_base


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    add_config_args(ap)
    ap.add_argument(
        "--write", action="store_true",
        help="Patch the upload sheets in place, creates .bak; default is dry-run.",
    )
    ap.add_argument(
        "--record-id",
        required=True,
        metavar="RECORD_ID",
        help="Zenodo record ID (e.g. 12345678)",
    )
    ap.add_argument(
        "--zip-dir",
        type=Path,
        default=None,
        help="Directory containing .zip files to process (default: <project>/Zenodo_upload/)",
    )
    ap.add_argument(
        "--metadata-xlsx",
        metavar="XLSX",
        help="Path to All-Metadata workbook (default: newest previous_metadata/*All*.xlsx)",
    )
    args = ap.parse_args()

    try:
        cfg = config_from_args(args)
    except ProjectRootError as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(2)

    url_template = f"https://zenodo.org/records/{args.record_id}/files/{{zipname}}?download=1&preview=1"

    zip_dir = args.zip_dir if args.zip_dir else (cfg.root / "Zenodo_upload")
    meta = Path(args.metadata_xlsx).resolve() if args.metadata_xlsx else cfg.master_workbook
    if meta is None or not meta.exists():
        print("ERROR: no metadata xlsx found. Pass --metadata-xlsx <path>.", file=sys.stderr)
        sys.exit(1)
    idx = load_curation_index(meta)
    print(f"Loaded {len(idx)} curated filenames from {meta.name}\n")

    sheet_map = discover_sheet_map(cfg.assay_sheets)
    if not sheet_map:
        print(f"WARNING: no upload sheets found in {cfg.assay_sheets}", file=sys.stderr)

    # For each zip: collect (UID, URL) pairs
    updates_by_sheet: dict[Path, list[tuple[str, str, str, str]]] = defaultdict(list)
    # (uid, url, source_filename, zip_name)

    for zip_path in sorted(zip_dir.glob("*.zip")):
        url = url_template.format(zipname=zip_path.name)
        print(f"--- {zip_path.name} ---")
        print(f"  URL: {url}")
        with zipfile.ZipFile(zip_path) as zf:
            for member in zf.namelist():
                if member.endswith("/"):
                    continue  # directory entry
                fname = Path(member).name
                entry = idx.get(fname)
                if not entry:
                    print(f"  WARNING: no curation match for {fname}")
                    continue
                stype, uid = entry
                sheet = sheet_map.get(stype)
                if not sheet:
                    print(f"  WARNING: no sheet mapping for {stype} (file {fname})")
                    continue
                if not sheet.exists():
                    print(f"  WARNING: sheet missing on disk: {sheet}")
                    continue
                updates_by_sheet[sheet].append((uid, url, fname, zip_path.name))
                print(f"    {uid:25s} <- {fname}")

    # Apply updates per sheet
    print("\n=== Applying updates ===")
    total_updated = 0
    for sheet_path, updates in updates_by_sheet.items():
        print(f"\n{sheet_path.name}: {len(updates)} updates")
        if not args.write:
            continue
        wb = load_workbook(sheet_path)
        try:
            ws = wb["Samples"]
            hdr = [c.value for c in ws[1]]
            # Add Link_PrimaryData column if missing
            if "Link_PrimaryData" not in hdr:
                ws.cell(row=1, column=len(hdr) + 1).value = "Link_PrimaryData"
                hdr.append("Link_PrimaryData")
                print(f"  added Link_PrimaryData column")
            try:
                i_uid = hdr.index("UID") + 1
            except ValueError:
                print(f"  SKIP {sheet_path.name}: no UID column")
                continue
            i_link = hdr.index("Link_PrimaryData") + 1
            # Build {uid → url} for this sheet's updates
            uid_to_url = {u: url for u, url, _, _ in updates}
            applied = 0
            for row in range(2, ws.max_row + 1):
                uid = ws.cell(row=row, column=i_uid).value
                if uid in uid_to_url:
                    ws.cell(row=row, column=i_link).value = uid_to_url[uid]
                    applied += 1
            shutil.copy(sheet_path, sheet_path.with_suffix(sheet_path.suffix + ".bak"))
            wb.save(sheet_path)
            print(f"  wrote {applied} Link_PrimaryData values")
            total_updated += applied
        finally:
            wb.close()

    if not args.write:
        print("\nDry run — no sheets modified. Re-run with --write to apply.")
    else:
        print(f"\nDone. {total_updated} rows updated across {len(updates_by_sheet)} sheets.")


if __name__ == "__main__":
    main()
