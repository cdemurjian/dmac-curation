#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = ["smbprotocol>=1.16", "python-dotenv>=1.0"]
# ///
"""Stream-pull files from an SMB share with on-the-fly gzip compression.

Reads source paths from a manifest (--from-manifest) or builds a manifest
by planning against the live SMB share. Streams each file via smbprotocol,
pipes through `pigz -c`, and writes gzipped output to the local destination.
Resumable (--resume), supports row slicing (--rows N-M or --rows-from FILE),
and is dry-run by default (pass --write to actually transfer).

Credentials read from .env via python-dotenv:
  MIT_USER (e.g. cdemu@mit.edu - domain embedded in the username)
  MIT_PASS
  SMB_HOST (default: bmc-pub14.mit.edu)
  SMB_SHARE (default: depends on lab)
VPN required for MIT BMC server.

Usage:
  uv run scripts/smb_pull.py                  # build manifest + size estimate, no transfer (default)
  uv run scripts/smb_pull.py --write          # actually pull; streams SMB -> pigz -> .fastq.gz
  uv run scripts/smb_pull.py --write --resume # skip already-completed outputs
  uv run scripts/smb_pull.py --write --from-manifest  # skip planning, read jobs from manifest.tsv
  uv run scripts/smb_pull.py --batch 200710Eng        # one batch only, preview

Originally consolidated from pull_spatial.py + pull_bulk_rna.py in the
srp/lee curation session. The sample-planning functions (load_manuscript_samples,
plan_sample, PLATE_STRATEGY) retain lee-specific structure — see TODO comments
for v0.2 generalisation. The --from-manifest path is fully generic.
"""
from __future__ import annotations

import argparse
import csv
import os
import subprocess
import sys
import time
from collections import defaultdict
from pathlib import Path

from dotenv import load_dotenv

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _config import add_config_args, config_from_args  # noqa: E402
from _config import ProjectRootError  # noqa: E402

_PLUGIN = Path(__file__).resolve().parents[1]
# cwd .env first, plugin .env second — matches fdh_api.py:161.
for _candidate in (Path.cwd() / ".env", _PLUGIN / ".env"):
    if _candidate.exists():
        load_dotenv(_candidate, override=False)

import smbclient  # noqa: E402 — must follow load_dotenv

HOST = os.environ.get("SMB_HOST", "bmc-pub14.mit.edu")
SHARE = os.environ.get("SMB_SHARE")  # no default share name; validated in main()
MIT_USER = os.environ.get("MIT_USER", "")
MIT_PASS = os.environ.get("MIT_PASS", "")

# ---------------------------------------------------------------------------
# Lee/Engelward-lab plate strategy
# TODO(v0.2): move this table out into a TOML/JSON config so smb_pull.py is
# truly project-agnostic. The plate names and batch IDs below are specific to
# the srp/lee bulk RNA-seq curation session.
# ---------------------------------------------------------------------------
PLATE_STRATEGY = {
    "191212Eng": {"plates": ["4348L", "4353L"], "merge": True,  "year_prefix": "D19"},
    "200710Eng": {"plates": ["4534L"],          "merge": False, "year_prefix": "D20"},
    "200724Eng": {"plates": ["4569G"],          "merge": False, "year_prefix": "D20"},
    "200803Eng": {"plates": ["4580L", "4584L"], "merge": True,  "year_prefix": "D20"},
    "210901Eng": {"plates": ["5342L"],          "merge": False, "year_prefix": "D21"},
}

CHUNK = 1 << 20  # 1 MB read chunks from SMB


# ---------------------------------------------------------------------------
# Sample loading (lee/Engelward-specific — requires the manuscript xlsx)
# TODO(v0.2): replace load_manuscript_samples() with a generic CSV/TSV loader
# so this path works without openpyxl and without a project-specific workbook.
# ---------------------------------------------------------------------------
def load_manuscript_samples(master_workbook: Path | None) -> list[dict]:
    """Read manuscript bulk RNA samples from the project master workbook.

    NOTE: This function is srp/lee-specific. It expects a workbook with columns
    'Sample Annotation', 'Mouse Acc.# ID#', 'Folder Name' — resolved from the
    project's previous_metadata/*All*.xlsx (or --master-baseline).
    TODO(v0.2): generalise — accept a --sample-list CSV instead of openpyxl.
    """
    if master_workbook is None:
        sys.exit(
            "ERROR: no master workbook found for sample planning. Add one to\n"
            "previous_metadata/*All*.xlsx, pass --master-baseline <path>, or use\n"
            "--from-manifest to skip planning entirely."
        )
    try:
        from openpyxl import load_workbook  # noqa: PLC0415 — lazy, not in PEP 723 deps
    except ImportError:
        sys.exit(
            "openpyxl is required for sample planning but is not installed.\n"
            "Either add openpyxl to your environment or use --from-manifest to\n"
            "skip planning entirely."
        )
    wb = load_workbook(master_workbook, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])  # TODO(v0.2): unguarded — raises IndexError on empty workbook; add check
    i_sa = hdr.index("Sample Annotation")
    i_acc = hdr.index("Mouse Acc.# ID#")
    i_folder = hdr.index("Folder Name")
    out = []
    for r in rows[1:]:
        if not any(c is not None for c in r):
            continue
        out.append({
            "sample_annotation": str(r[i_sa]).strip() if r[i_sa] else "",
            "mouse_acc":         str(r[i_acc]).strip() if r[i_acc] else "",
            "folder":            str(r[i_folder]).strip() if r[i_folder] else "",
        })
    return out


def detect_files(sample_dir: str) -> dict:
    """List FASTQs in a server per-sample dir, classify by *_1_/_2_/_NA_ pattern."""
    fastqs: dict[str, list] = {"R1": [], "R2": [], "NA": []}
    for e in smbclient.scandir(sample_dir):
        if e.is_dir():
            continue
        if "_sequence.fastq" not in e.name:
            continue
        sz = e.stat().st_size
        path = rf"{sample_dir}\{e.name}"
        if "_1_sequence.fastq" in e.name:
            fastqs["R1"].append((path, sz, e.name))
        elif "_2_sequence.fastq" in e.name:
            fastqs["R2"].append((path, sz, e.name))
        elif "_NA_sequence.fastq" in e.name:
            fastqs["NA"].append((path, sz, e.name))
    return fastqs


def plan_sample(sample: dict, out_dir: Path) -> list[dict]:
    """Build the pull plan for one manuscript sample.

    Returns a list of job dicts. Each job is one local .fastq.gz output.
    For single-end samples merged across plates: 1 job.
    For paired-end: 2 jobs (R1 and R2; each may span multiple plates).

    TODO(v0.2): the server path template (users/<user>/<batch>/<id>-<plate>)
    is srp/lee-specific. Parameterise via --server-path-template.
    """
    folder = sample["folder"]
    sa = sample["sample_annotation"]
    strat = PLATE_STRATEGY.get(folder)
    if not strat:
        return [{"error": f"no plate strategy for batch {folder}"}]
    server_sample_id = f"{strat['year_prefix']}-{sa[1:]}"  # e.g. D389001 -> D19-389001

    # Discover files in each plate dir
    # TODO(v0.2): 'noraho' below is a srp/lee user path — parameterise via --server-user or template
    plate_files: dict[str, dict] = {}
    for plate in strat["plates"]:
        sample_dir = rf"\\{HOST}\{SHARE}\users\noraho\{folder}\{server_sample_id}-{plate}"
        try:
            plate_files[plate] = detect_files(sample_dir)
        except Exception as ex:
            return [{"error": f"{server_sample_id}-{plate}: {ex}"}]

    has_paired = any(plate_files[p]["R1"] for p in strat["plates"])
    has_single = any(plate_files[p]["NA"] for p in strat["plates"])

    out_subdir = out_dir / folder
    jobs: list[dict] = []
    if has_paired and not has_single:
        for stream in ("R1", "R2"):
            sources = []
            for plate in strat["plates"]:
                sources.extend(plate_files[plate][stream])
            if not sources:
                jobs.append({"error": f"{sa}: missing {stream} stream"})
                continue
            jobs.append({
                "sample_annotation": sa,
                "batch": folder,
                "output": str(out_subdir / f"{folder}_{sa}_{stream}.fastq.gz"),
                "sources": [(p, sz, n) for (p, sz, n) in sources],
                "kind": "paired-stream",
                "stream": stream,
            })
    elif has_single and not has_paired:
        sources = []
        for plate in strat["plates"]:
            sources.extend(plate_files[plate]["NA"])
        if not sources:
            jobs.append({"error": f"{sa}: missing _NA_ FASTQ"})
        else:
            jobs.append({
                "sample_annotation": sa,
                "batch": folder,
                "output": str(out_subdir / f"{folder}_{sa}.fastq.gz"),
                "sources": [(p, sz, n) for (p, sz, n) in sources],
                "kind": "single" if len(sources) == 1 else "single-merged",
                "stream": "NA",
            })
    elif has_paired and has_single:
        jobs.append({"error": f"{sa}: AMBIGUOUS — has both paired and _NA_ files in same plate dirs"})
    else:
        jobs.append({"error": f"{sa}: no FASTQs found in plate dirs {strat['plates']}"})
    return jobs


# ---------------------------------------------------------------------------
# Manifest I/O
# ---------------------------------------------------------------------------

def write_manifest(jobs: list[dict], manifest: Path):
    manifest.parent.mkdir(parents=True, exist_ok=True)
    with open(manifest, "w") as fh:
        w = csv.writer(fh, delimiter="\t")
        w.writerow([
            "sample_annotation", "batch", "kind", "stream",
            "total_src_bytes", "n_sources", "output",
            "source_paths", "source_sizes", "error",
        ])
        for j in jobs:
            if "error" in j:
                w.writerow(["", "", "", "", "", "", "", "", "", j["error"]])
            else:
                total = sum(sz for _, sz, _ in j["sources"])
                w.writerow([
                    j["sample_annotation"], j["batch"], j["kind"], j["stream"],
                    total, len(j["sources"]),
                    j["output"],
                    ";".join(p for p, _, _ in j["sources"]),
                    ";".join(str(sz) for _, sz, _ in j["sources"]),
                    "",
                ])
    print(f"Manifest written to {manifest}")


def load_jobs_from_manifest(manifest: Path) -> list[dict]:
    """Read jobs directly from manifest.tsv (no SMB planning required)."""
    jobs = []
    with open(manifest) as fh:
        r = csv.DictReader(fh, delimiter="\t")
        for row in r:
            if row.get("error"):
                continue
            paths = row["source_paths"].split(";")
            sizes = [int(s) for s in row["source_sizes"].split(";")]
            sources = [
                (p, sz, Path(p.replace("\\\\", "/").replace("\\", "/")).name)
                for p, sz in zip(paths, sizes)
            ]
            jobs.append({
                "sample_annotation": row["sample_annotation"],
                "batch":  row["batch"],
                "kind":   row["kind"],
                "stream": row["stream"],
                "output": row["output"],
                "sources": sources,
            })
    return jobs


# ---------------------------------------------------------------------------
# Pull engine
# ---------------------------------------------------------------------------

def do_pull(job: dict, resume: bool):
    """Stream one job from SMB through pigz to a local .fastq.gz file."""
    out = Path(job["output"])
    if resume and out.exists() and out.stat().st_size > 0:
        print(f"  SKIP (exists): {out}")
        return
    out.parent.mkdir(parents=True, exist_ok=True)
    tmp = out.with_suffix(out.suffix + ".partial")
    t0 = time.time()
    total_in = sum(sz for _, sz, _ in job["sources"])  # TODO(v0.2): dead code, remove — total_in is never read
    bytes_in = 0
    with open(tmp, "wb") as fh_out:
        proc = subprocess.Popen(["pigz", "-c"], stdin=subprocess.PIPE, stdout=fh_out)
        try:
            for src_path, sz, name in job["sources"]:
                with smbclient.open_file(src_path, mode="rb", share_access="rwd") as fh_in:
                    while True:
                        chunk = fh_in.read(CHUNK)
                        if not chunk:
                            break
                        proc.stdin.write(chunk)
                        bytes_in += len(chunk)
            proc.stdin.close()
            ret = proc.wait()
            if ret != 0:
                raise RuntimeError(f"pigz exited {ret}")
        except Exception:
            proc.kill()
            tmp.unlink(missing_ok=True)
            raise
    tmp.rename(out)
    out_sz = out.stat().st_size
    dur = time.time() - t0
    rate = bytes_in / 1e6 / max(dur, 1e-3)
    print(
        f"  OK {out.name}  {bytes_in/1e9:.2f} GB -> {out_sz/1e9:.2f} GB"
        f"  ({100*out_sz/max(bytes_in, 1):.1f}%)  {dur:.1f}s  {rate:.0f} MB/s"
    )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    add_config_args(ap)
    ap.add_argument(
        "--write", action="store_true",
        help="Perform the transfer; default is dry-run "
             "(manifest + size estimate only).",
    )
    ap.add_argument(
        "--out-dir", type=Path, default=None,
        help="Destination for pulled files "
             "(default: <project-root>/GEO/bulk_rna/fastq)",
    )
    ap.add_argument(
        "--manifest", type=Path, default=None,
        help="Manifest TSV path "
             "(default: <project-root>/GEO/bulk_rna/manifest.tsv)",
    )
    ap.add_argument(
        "--batch",
        help="restrict planning to one batch (e.g. 200710Eng) — no effect with --from-manifest",
    )
    ap.add_argument(
        "--resume", action="store_true",
        help="skip output files that already exist and are non-empty",
    )
    ap.add_argument(
        "--from-manifest", action="store_true",
        help="skip SMB planning; read jobs from existing manifest.tsv",
    )
    ap.add_argument(
        "--rows",
        help="filter to job rows N-M (1-indexed inclusive, e.g. 1-107)",
    )
    ap.add_argument(
        "--rows-from",
        metavar="FILE",
        help="file with one row number per line; filter to those job rows",
    )
    args = ap.parse_args()

    try:
        cfg = config_from_args(args)
    except ProjectRootError as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(2)

    out_dir = args.out_dir if args.out_dir else (cfg.root / "GEO" / "bulk_rna" / "fastq")
    manifest = args.manifest if args.manifest else (cfg.root / "GEO" / "bulk_rna" / "manifest.tsv")

    # Credentials check — fail fast with a clear message rather than an obscure KeyError
    if not MIT_USER or not MIT_PASS:
        sys.exit(
            "Credentials not set. Ensure MIT_USER and MIT_PASS are defined in .env or the environment.\n"
            "VPN is required for the MIT BMC SMB server."
        )

    # Share name is required — no hardcoded lab default.
    if not SHARE:
        print("ERROR: set SMB_SHARE in .env (no default share name)", file=sys.stderr)
        sys.exit(2)

    # Verify pigz is available before any SMB work — fail fast on a missing tool
    if args.write and subprocess.call(
        ["which", "pigz"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
    ) != 0:
        sys.exit("pigz not found; install it (e.g. sudo dnf install pigz) and retry.")

    by_batch_stats: dict = defaultdict(lambda: {"jobs": 0, "src_bytes": 0, "errors": 0})

    if args.from_manifest:
        # --from-manifest: read the plan locally; skip SMB planning entirely.
        # register_session is deferred until we actually start pulling below.
        if not manifest.exists():
            sys.exit(f"--from-manifest requires {manifest} to exist. Run without --write first.")
        jobs_all = load_jobs_from_manifest(manifest)
        print(f"Loaded {len(jobs_all)} jobs from manifest (planning skipped)")
        for j in jobs_all:
            b = j["batch"]
            by_batch_stats[b]["jobs"] += 1
            by_batch_stats[b]["src_bytes"] += sum(sz for _, sz, _ in j["sources"])
    else:
        # Planning phase requires SMB connectivity — register session now.
        smbclient.register_session(HOST, username=MIT_USER, password=MIT_PASS)

        samples = load_manuscript_samples(cfg.master_workbook)
        if args.batch:
            samples = [s for s in samples if s["folder"] == args.batch]
        print(f"Samples (filtered): {len(samples)}")

        jobs_all = []
        print("Planning (one SMB round-trip per plate dir)...")
        for i, s in enumerate(samples, 1):
            if i == 1 or i % 10 == 0 or i == len(samples):
                print(f"  [{i:>3d}/{len(samples)}] {s['folder']} {s['sample_annotation']}", flush=True)
            plan = plan_sample(s, out_dir)
            for j in plan:
                jobs_all.append(j)
                b = s["folder"]
                if "error" in j:
                    by_batch_stats[b]["errors"] += 1
                else:
                    by_batch_stats[b]["jobs"] += 1
                    by_batch_stats[b]["src_bytes"] += sum(sz for _, sz, _ in j["sources"])

    # Write the full manifest BEFORE any row-filtering so the manifest always
    # reflects the complete plan.  (Finding 1: filtering mutated jobs_all in place
    # and wrote only the slice — destroying the full manifest.)
    if not args.from_manifest:
        write_manifest(jobs_all, manifest)

    # --rows N-M filter — applies only to the pull phase, not to the manifest
    jobs_to_pull = list(jobs_all)
    if args.rows:
        try:
            lo, hi = (int(x) for x in args.rows.split("-"))
        except ValueError:
            sys.exit(f"--rows must be like '1-107', got {args.rows!r}")
        original = len(jobs_to_pull)
        jobs_to_pull = jobs_to_pull[lo - 1:hi]
        print(f"Row filter {args.rows}: {len(jobs_to_pull)}/{original} jobs selected")

    # --rows-from FILE filter
    elif args.rows_from:
        # TODO(v0.2): bare ValueError here if a line isn't an integer — add proper error message
        wanted = {int(x) for x in Path(args.rows_from).read_text().split() if x.strip()}
        original = len(jobs_to_pull)
        jobs_to_pull = [j for i, j in enumerate(jobs_to_pull, 1) if i in wanted]
        print(f"Row-list filter ({args.rows_from}): {len(jobs_to_pull)}/{original} jobs selected")

    print("\n=== Per-batch summary ===")
    grand_bytes = 0
    grand_jobs = 0
    grand_errs = 0
    for b, st in sorted(by_batch_stats.items()):
        print(f"  {b}: {st['jobs']} jobs, {st['src_bytes']/1e9:.1f} GB source, {st['errors']} errors")
        grand_bytes += st["src_bytes"]
        grand_jobs += st["jobs"]
        grand_errs += st["errors"]
    print(
        f"  TOTAL: {grand_jobs} jobs, {grand_bytes/1e9:.1f} GB source uncompressed, "
        f"~{grand_bytes * 0.3 / 1e9:.0f} GB gzipped est.   errors: {grand_errs}"
    )

    errs = [j for j in jobs_to_pull if "error" in j]
    if errs:
        print(f"\n{len(errs)} errors (first 10):")
        for j in errs[:10]:
            print(f"  {j['error']}")

    if not args.write:
        print("\nDry run only — no files transferred. Inspect manifest.tsv before re-running with --write.")
        return

    if errs:
        print(f"\nABORTING: {len(errs)} errors found in plan. Resolve before pulling.")
        sys.exit(1)

    # For --from-manifest: session not yet registered; do it now before pulling.
    if args.from_manifest:
        smbclient.register_session(HOST, username=MIT_USER, password=MIT_PASS)

    # jobs_to_pull denominator: only non-error rows in the (possibly filtered) pull set
    total_jobs_to_pull = sum(1 for j in jobs_to_pull if "error" not in j)
    print(f"\nPulling {total_jobs_to_pull} jobs ({grand_bytes/1e9:.1f} GB uncompressed)...")
    pull_idx = 0
    for j in jobs_to_pull:
        if "error" in j:
            continue
        pull_idx += 1
        print(f"[{pull_idx}/{total_jobs_to_pull}] {j['sample_annotation']} {j['stream']}  ({sum(sz for _, sz, _ in j['sources'])/1e9:.2f} GB)")
        do_pull(j, resume=args.resume)


if __name__ == "__main__":
    main()
