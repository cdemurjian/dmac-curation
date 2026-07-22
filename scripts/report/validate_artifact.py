# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Validate a rendered submission artifact.

A format is not "supported" until it has a renderer AND a validator. This is the
validator half.

VENDORED, IN PART, from tavjo/dmac-assistant `tools/hibayes/artifact_validator.py`.
Upstream is 897 lines and imports `tools.hibayes.enums` and
`tools.hibayes.exporter`; most of it is a 29-column evidence-CSV harness this
plugin has no use for. Taken here: the ArtifactStatus enum, ValidatorResult, the
GEO required-field logic, and the read-only-workbook single-pass discipline.
Added here: SRA (two sections) and PRIDE (tab-delimited, not a spreadsheet).
See context/PROVENANCE.json.

The valuable part of upstream's GEO check, preserved verbatim in spirit: TWO
independent checks, not one. (i) each single-`*`-prefixed required field is
present as a COLUMN HEADER, and (ii) each required column is NON-NULL on every
data row. An earlier upstream version flattened every cell into one string and
substring-scanned for field names, which conflated header presence with an
arbitrary mention -- `*title` appearing in a free-form study cell falsely
satisfied "present".
"""
from __future__ import annotations

import json
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any


class ArtifactStatus(str, Enum):
    Valid = "Valid"
    Incomplete = "Incomplete"
    SchemaInvalid = "SchemaInvalid"
    Missing = "Missing"
    Unreadable = "Unreadable"


# Maps onto the pipeline's QA vocabulary (PHASES.md Phase 9).
DISPOSITION = {
    ArtifactStatus.Valid: "CLEAN",
    ArtifactStatus.Incomplete: "SOFT_FLAG",
    ArtifactStatus.SchemaInvalid: "HARD_REJECT",
    ArtifactStatus.Missing: "HARD_REJECT",
    ArtifactStatus.Unreadable: "HARD_REJECT",
}


@dataclass
class ValidatorResult:
    """Per-artifact validator output."""

    status: ArtifactStatus
    parser_used: str | None = None
    parse_success: bool | None = None
    sheet_count: int | None = None
    row_count: int | None = None
    column_count: int | None = None
    nonempty_cell_count: int | None = None
    null_cell_fraction: float | None = None
    required_fields_present: bool | None = None
    required_fields_complete: bool | None = None
    missing_required_fields: str | None = None
    all_required_rows_complete: bool | None = None
    validation_notes: str = ""

    @property
    def disposition(self) -> str:
        return DISPOSITION[self.status]


def required_fields(spec_path: Path, section: str) -> list[str]:
    """Single-`*`-prefixed keys of `spec[section][0]`.

    A single `*` means required. A double `**` means conditionally required and
    is deliberately excluded: GEO's `**tissue` and `**cell line` are alternatives,
    and demanding both would reject every valid workbook.

    KNOWN GAP, verified: SRA's `libraries` section marks NOTHING with `*` -
    `sample_name`, `library_ID`, `library_strategy` and the rest are all bare.
    So this returns [] for it, and _validate_xlsx then reports Valid for any
    readable workbook. That is honest (the template asserts no requirements) but
    weak. `biosamples` DOES star its required fields, so SRA is not unguarded
    overall. If SRA library validation needs teeth, add an explicit required
    list here rather than inventing stars in the vendored template - the
    template's provenance entry says `local_divergence: none` and it should
    stay that way.
    """
    spec = json.loads(Path(spec_path).read_text())
    rows = spec.get(section)
    if not isinstance(rows, list) or not rows or not isinstance(rows[0], dict):
        raise SystemExit(
            f"{spec_path}: section {section!r} must be a non-empty list whose "
            f"first element is a dict"
        )
    return [k for k in rows[0]
            if k.startswith("*") and not k.startswith("**")]


def _read_sheet_once(ws, required: list[str]) -> tuple[list[str], list[tuple], int, int]:
    """Collect header, data rows and cell counts in ONE pass.

    In read_only mode the underlying archive closes with the workbook, and any
    later iter_rows on the worksheet raises. Every row is buffered in a single
    iteration before anything is computed.

    The header row is LOCATED, not assumed to be row 0. The real GEO template's
    `Metadata` sheet is a VERTICAL multi-section form: a `# ...` comment on row
    1, then a STUDY block, then the SAMPLES header (`*library name` and the rest)
    partway down, then a blank gap and a PROTOCOLS block. Assuming row 0 finds
    the comment, not the header, so none of the required sample fields are seen
    and a valid artifact is wrongly HARD_REJECTed. Instead:

      - If `required` is non-empty, the header is the row with the MOST cells
        matching a required field name (case-insensitive). If no row overlaps,
        fall back to row 0. For any flat single-header sheet this IS row 0.
      - If `required` is empty (e.g. SRA `libraries` stars nothing), fall back
        to row 0 - preserving the prior behavior for that path exactly.

    The DATA rows are the rows below the located header, stopping at the first
    ENTIRELY EMPTY row (all cells None/"") - so the GEO sample block ends at the
    blank gap and the trailing PROTOCOLS/STUDY rows are not read as sample data.
    Counts are over the located block (header row + data rows).
    """
    rows = [row for row in ws.iter_rows(values_only=True)]
    if not rows:
        return [], [], 0, 0

    header_idx = 0
    if required:
        wanted = {r.strip().lower() for r in required}
        best = 0
        for i, row in enumerate(rows):
            overlap = sum(1 for v in row
                          if v is not None and str(v).strip().lower() in wanted)
            if overlap > best:
                best, header_idx = overlap, i

    header = ["" if v is None else str(v) for v in rows[header_idx]]

    data: list[tuple[Any, ...]] = []
    for row in rows[header_idx + 1:]:
        if all(v in (None, "") for v in row):
            break
        data.append(row)

    nonempty = total = 0
    for row in (rows[header_idx], *data):
        for value in row:
            total += 1
            if value not in (None, ""):
                nonempty += 1
    return header, data, nonempty, total


def _check_required(header: list[str], data: list[tuple],
                    required: list[str]) -> tuple[bool, bool, str | None]:
    """(headers_present, rows_complete, missing_summary)."""
    header_lower = [h.strip().lower() for h in header]
    missing_headers = [r for r in required if r.lower() not in header_lower]
    present = not missing_headers

    rows_complete = True
    per_row_missing: list[str] = []
    for r in required:
        if r.lower() not in header_lower:
            continue
        col = header_lower.index(r.lower())
        for row in data:
            if col >= len(row):
                rows_complete = False
                per_row_missing.append(r)
                break
            cell = row[col]
            if cell is None or (isinstance(cell, str) and cell.strip() == ""):
                rows_complete = False
                per_row_missing.append(r)
                break

    all_missing = list(missing_headers) + per_row_missing
    return present, rows_complete, (";".join(all_missing) if all_missing else None)


def _validate_xlsx(file_path: Path, required: list[str],
                   parser_note: str) -> ValidatorResult:
    file_path = Path(file_path)
    if not file_path.exists():
        return ValidatorResult(status=ArtifactStatus.Missing,
                               validation_notes=f"file not at {file_path}")
    try:
        import openpyxl
    except ImportError:
        return ValidatorResult(status=ArtifactStatus.Unreadable,
                               parser_used="openpyxl", parse_success=False,
                               validation_notes="openpyxl not importable")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        return ValidatorResult(status=ArtifactStatus.Unreadable,
                               parser_used="openpyxl", parse_success=False,
                               validation_notes=f"openpyxl error: {type(exc).__name__}")

    sheet_count = len(wb.sheetnames)
    if sheet_count == 0:
        wb.close()
        return ValidatorResult(status=ArtifactStatus.SchemaInvalid,
                               parser_used="openpyxl", parse_success=True,
                               sheet_count=0,
                               validation_notes="workbook has zero sheets")

    ws = wb[wb.sheetnames[0]]
    header, data, nonempty, total = _read_sheet_once(ws, required)
    wb.close()

    # Structural counts describe the LOCATED block, not the raw sheet extent:
    # column_count is the header's width, row_count the header row plus its data
    # rows (so a vertical-form sheet reports its SAMPLES block, not the whole
    # multi-section form).
    column_count = len(header)
    row_count = (1 + len(data)) if header else 0

    present, rows_complete, missing = _check_required(header, data, required)
    if not present:
        status = ArtifactStatus.SchemaInvalid
    elif not rows_complete:
        status = ArtifactStatus.Incomplete
    else:
        status = ArtifactStatus.Valid

    return ValidatorResult(
        status=status,
        parser_used="openpyxl",
        parse_success=True,
        sheet_count=sheet_count,
        row_count=row_count,
        column_count=column_count,
        nonempty_cell_count=nonempty,
        null_cell_fraction=((total - nonempty) / total) if total else None,
        required_fields_present=present,
        required_fields_complete=present and rows_complete,
        missing_required_fields=missing,
        all_required_rows_complete=rows_complete if present else False,
        validation_notes=parser_note,
    )


def validate_geo_xlsx(*, file_path: Path,
                      geo_template_path: Path | None) -> ValidatorResult:
    """GEO `.xlsx`: structural counts plus the two-part required-field check."""
    required = (required_fields(geo_template_path, "samples")
                if geo_template_path and Path(geo_template_path).is_file() else [])
    return _validate_xlsx(file_path, required, "GEO samples section")


def validate_sra_xlsx(*, file_path: Path, sra_spec_path: Path,
                      section: str = "libraries") -> ValidatorResult:
    """SRA `.xlsx`. Two workbooks ship: `libraries` and `biosamples`."""
    required = (required_fields(sra_spec_path, section)
                if Path(sra_spec_path).is_file() else [])
    return _validate_xlsx(file_path, required, f"SRA {section} section")


def validate_pride_px(*, file_path: Path, pride_spec_path: Path) -> ValidatorResult:
    """PRIDE submission summary file - TAB-DELIMITED, not a spreadsheet.

    pride.json declares output_type 'tab-delimited PRIDE submission summary
    file' with line prefixes MTD / FMH / FME / SMH / SME / COM. chat_nextseek's
    e2e catalog asserts `pride.xlsx`, which is the wrong artifact type.
    """
    file_path = Path(file_path)
    if not file_path.exists():
        return ValidatorResult(status=ArtifactStatus.Missing,
                               validation_notes=f"file not at {file_path}")
    try:
        text = file_path.read_text(encoding="utf-8")
    except (OSError, UnicodeDecodeError) as exc:
        return ValidatorResult(status=ArtifactStatus.Unreadable,
                               parser_used="px-tsv", parse_success=False,
                               validation_notes=f"unreadable: {type(exc).__name__}")

    spec = json.loads(Path(pride_spec_path).read_text())
    allowed = set((spec["format"]["line_prefixes"]).values())

    lines = [l for l in text.splitlines() if l.strip()]
    seen: dict[str, int] = {}
    bad: list[str] = []
    nonempty = 0
    for line in lines:
        prefix = line.split("\t", 1)[0]
        if prefix not in allowed:
            bad.append(prefix)
            continue
        seen[prefix] = seen.get(prefix, 0) + 1
        nonempty += sum(1 for f in line.split("\t") if f.strip())

    notes = []
    if bad:
        notes.append(f"unknown line prefixes: {sorted(set(bad))[:5]}")
    if not seen.get("MTD"):
        notes.append("no MTD (metadata) lines")
    if not seen.get("SME"):
        notes.append("no SME (sample metadata) rows")

    if bad or not seen.get("MTD"):
        status = ArtifactStatus.SchemaInvalid
    elif not seen.get("SME"):
        status = ArtifactStatus.Incomplete
    else:
        status = ArtifactStatus.Valid

    return ValidatorResult(
        status=status,
        parser_used="px-tsv",
        parse_success=True,
        sheet_count=1,
        row_count=len(lines),
        column_count=max((len(l.split("\t")) for l in lines), default=0),
        nonempty_cell_count=nonempty,
        required_fields_present=bool(seen.get("MTD")),
        all_required_rows_complete=bool(seen.get("SME")),
        validation_notes="; ".join(notes) or "px line prefixes valid",
    )
