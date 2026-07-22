# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Render a filled report document to its submission artifact.

A real dispatcher, deliberately: chat_nextseek's reports/outputs.py is a
400-line function with a hardcoded if/elif format dispatch, and porting that
shape is an explicit non-goal.

Each format's artifact is what its own spec says it is:

  GEO   -> one xlsx, via scripts/deposit/geo_build_xlsx.py (ours, kept)
  SRA   -> TWO xlsx: SRA_metadata (libraries) and SRA_biosample (biosamples)
  PRIDE -> submission.px, TAB-DELIMITED with MTD/FMH/FME/SMH/SME line prefixes.
           NOT a spreadsheet. chat_nextseek's e2e catalog asserts pride.xlsx,
           which is the wrong artifact type, and it has no exporter at all.
"""
from __future__ import annotations

import json
import subprocess
from pathlib import Path

from openpyxl import load_workbook

_PLUGIN = Path(__file__).resolve().parent.parent.parent


class UnsupportedFormatError(Exception):
    """No renderer for this format. A format is not supported without one."""


def _strip_marker(name: str) -> str:
    """`*species` / `**tissue` -> `species` / `tissue`.

    The stars are the template's required/conditional markers, not wire names.
    """
    return name.lstrip("*")


# ── GEO ────────────────────────────────────────────────────────────────────

def render_geo(filled: dict, template_xlsx: Path, out_path: Path) -> Path:
    """Delegate to scripts/deposit/geo_build_xlsx.py - ours, kept.

    Two implementations existed and differed; the spec's instruction is to pick
    one and delete the other. Ours wins because it is already a PEP 723 uv
    script, cwd-relative and arg-driven. Its input contract - `data["samples"]`
    plus `data.get("paired_end_experiments", [])` - is already exactly what
    execute.apply_mapping emits for GEO.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_json = out_path.with_suffix(".render-input.json")
    tmp_json.write_text(json.dumps(filled, indent=2))
    script = _PLUGIN / "scripts" / "deposit" / "geo_build_xlsx.py"
    result = subprocess.run(
        ["uv", "run", "--script", str(script),
         str(tmp_json), str(template_xlsx), str(out_path)],
        capture_output=True, text=True, timeout=300,
    )
    if result.returncode != 0:
        raise RuntimeError(f"geo_build_xlsx failed:\n{result.stderr}")
    tmp_json.unlink(missing_ok=True)
    return out_path


# ── SRA ────────────────────────────────────────────────────────────────────

def _find_header(wb, record_keys: set[str], scan_rows: int = 50):
    """Locate the worksheet + row whose headers best match `record_keys`.

    Deliberately NOT "the first row with >= N non-empty cells": the vendored
    SRA templates lead with rows that would defeat that heuristic. SRA_metadata's
    first sheet is a "Contact Info and Instructions" sheet whose only multi-cell
    row is instruction labels; the real `sample_name`/`library_ID` header lives
    in row 1 of the *second* sheet (SRA_data). SRA_biosample's first eleven rows
    are single-cell `#` comment lines and the header is row 12. Matching the row
    that actually contains the known column names finds the right sheet and row
    in both, and is insensitive to the template's `*`/`#` markers.

    Returns `(match_count, worksheet, header_row_idx, headers)` or None.
    """
    wanted = {_strip_marker(str(k)).lower() for k in record_keys}
    best = None
    for ws in wb.worksheets:
        for i, row in enumerate(ws.iter_rows(values_only=True, max_row=scan_rows),
                                start=1):
            headers = ["" if c is None else str(c).strip() for c in row]
            names = {_strip_marker(h).lower() for h in headers if h}
            match = len(wanted & names)
            if match and (best is None or match > best[0]):
                best = (match, ws, i, headers)
    return best


def _fill_sheet_from_rows(template: Path, rows: list[dict], out_path: Path) -> Path:
    """Write `rows` under the template's own header row, matched by name.

    Header matching is case-insensitive and marker-insensitive, so a template
    header `sample_name` accepts a filled key `*sample_name`. The matched sheet
    is promoted to the front so the rendered artifact leads with its data (NCBI
    keeps SRA_data on a later tab; nothing is dropped, only reordered).
    """
    wb = load_workbook(template)
    try:
        keys = {k for record in rows for k in record}
        found = _find_header(wb, keys)
        if found is None:
            raise ValueError(
                f"no header row matching {sorted(keys)} in {template}")
        _, ws, header_row_idx, headers = found
        wb.move_sheet(ws, offset=-wb.index(ws))

        lookup = {_strip_marker(h).lower(): i
                  for i, h in enumerate(headers, start=1) if h}
        for offset, record in enumerate(rows, start=1):
            for key, value in record.items():
                col = lookup.get(_strip_marker(str(key)).lower())
                if col:
                    ws.cell(row=header_row_idx + offset, column=col).value = value
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
    finally:
        wb.close()
    return out_path


def render_sra(filled: dict, metadata_xlsx: Path, biosample_xlsx: Path,
               out_dir: Path) -> list[Path]:
    """SRA ships TWO templates and therefore produces two workbooks."""
    out_dir = Path(out_dir)
    written = []
    if filled.get("libraries"):
        written.append(_fill_sheet_from_rows(
            metadata_xlsx, filled["libraries"],
            out_dir / "SRA_metadata_filled.xlsx"))
    if filled.get("biosamples"):
        written.append(_fill_sheet_from_rows(
            biosample_xlsx, filled["biosamples"],
            out_dir / "SRA_biosample_filled.xlsx"))
    return written


# ── PRIDE ──────────────────────────────────────────────────────────────────

def render_pride(filled: dict, spec_path: Path, out_path: Path) -> Path:
    """ProteomeXchange Submission Summary File, v2.2.1. Tab-delimited.

    Written from scratch: chat_nextseek has a pride.json template and a row-key
    entry but NO exporter, so it silently yields JSON only while its own e2e
    catalog asserts pride.xlsx. That assertion names the wrong artifact type.
    """
    spec = json.loads(Path(spec_path).read_text())
    prefixes = spec["format"]["line_prefixes"]
    lines: list[str] = []

    lines.append("\t".join([prefixes["comment"],
                            "Generated by dmac-curation report mode"]))
    for key, value in (filled.get("project_metadata") or {}).items():
        lines.append("\t".join([prefixes["metadata"], _strip_marker(str(key)),
                                "" if value is None else str(value)]))

    file_rows = filled.get("file_mapping") or []
    if file_rows:
        headers = [_strip_marker(k) for k in file_rows[0]]
        lines.append("\t".join([prefixes["file_mapping_header"], *headers]))
        for record in file_rows:
            lines.append("\t".join(
                [prefixes["file_mapping_entry"],
                 *["" if v is None else str(v) for v in record.values()]]))

    sample_rows = filled.get("sample_metadata") or []
    if sample_rows:
        headers = [_strip_marker(k) for k in sample_rows[0]]
        lines.append("\t".join([prefixes["sample_metadata_header"], *headers]))
        for record in sample_rows:
            lines.append("\t".join(
                [prefixes["sample_metadata_entry"],
                 *["" if v is None else str(v) for v in record.values()]]))

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return out_path


# ── dispatch ───────────────────────────────────────────────────────────────

def _geo(filled, template_dir: Path, out_dir: Path) -> list[Path]:
    return [render_geo(filled, template_dir / "GEO_template.xlsx",
                       out_dir / "GEO_filled.xlsx")]


def _sra(filled, template_dir: Path, out_dir: Path) -> list[Path]:
    return render_sra(filled, template_dir / "SRA_metadata.xlsx",
                      template_dir / "SRA_biosample.xlsx", out_dir)


def _pride(filled, template_dir: Path, out_dir: Path) -> list[Path]:
    return [render_pride(filled, template_dir / "pride.json",
                         out_dir / "submission.px")]


RENDERERS = {"GEO": _geo, "SRA": _sra, "PRIDE": _pride}


def render(report_type: str, filled: dict, *, template_dir: Path,
           out_dir: Path) -> list[Path]:
    """Render to every artifact the format requires. Returns the paths written."""
    key = str(report_type).upper()
    renderer = RENDERERS.get(key)
    if renderer is None:
        raise UnsupportedFormatError(
            f"no renderer for {report_type!r}; have {sorted(RENDERERS)}. "
            f"A format is not supported until it has a renderer AND a validator."
        )
    return renderer(filled, Path(template_dir), Path(out_dir))
