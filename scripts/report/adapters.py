# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Normalize whatever metadata the curator has into one internal shape.

Inputs are NOT a mode switch. Each adapter emits the same structure, and every
downstream step - protocol resolution, the LLM mapping step, rendering,
validation - is adapter-agnostic and sees only that structure.

    {"samples": [{"sample_type": "D.SEQ",
                  "uid": "D.SEQ-...",
                  "metadata": {<flat key/value>},
                  "parent": "TIS-..."}]}

Lineage is the flat `Parent` key, an upward UID pointer - NOT nesting. The
NExtSEEK retrieve response is nested five levels
(`data.data[i].samples[j].metadata`) and this module flattens it.

Enrichment is additive, never required. An adapter with no fetcher simply
returns no API-sourced samples rather than failing.
"""
from __future__ import annotations

import csv
import json
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

_MAX_LINEAGE_DEPTH = 12


@dataclass
class NormalizedSample:
    sample_type: str
    uid: str
    metadata: dict
    parent: str | None = None


@dataclass
class NormalizedInput:
    samples: list[NormalizedSample] = field(default_factory=list)
    source: dict = field(default_factory=dict)


def _sample_type_from_uid(uid: str) -> str:
    return str(uid).split("-", 1)[0] if uid and "-" in str(uid) else ""


def _clean(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


# ── adapters ───────────────────────────────────────────────────────────────

def adapt_uids(uids: list[str], *, fetch=None) -> NormalizedInput:
    """UIDs -> POST /nextseek_api/admin/samples/retrieve/.

    `fetch` is a callable taking a UID list and returning the parsed response,
    so this is testable without a network. With no fetcher, returns no samples:
    enrichment is additive, never required.
    """
    out = NormalizedInput(source={"adapter": "uids", "uids": list(uids)})
    if fetch is None:
        return out
    payload = fetch(list(uids)) or {}
    for group in ((payload.get("data") or {}).get("data") or []):
        for entry in (group.get("samples") or []):
            meta = dict(entry.get("metadata") or {})
            uid = _clean(meta.get("UID")) or ""
            out.samples.append(NormalizedSample(
                sample_type=_clean(meta.get("SampleType")) or _sample_type_from_uid(uid),
                uid=uid,
                metadata=meta,
                parent=_clean(meta.get("Parent")),
            ))
    return out


def adapt_retrieve_txt(path: Path, *, fetch=None) -> NormalizedInput:
    uids = [l.strip() for l in Path(path).read_text().splitlines() if l.strip()]
    out = adapt_uids(uids, fetch=fetch)
    out.source = {"adapter": "retrieve_txt", "path": str(path), "uids": uids}
    return out


def _rows_from_sheet(ws) -> list[dict]:
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []
    names = [str(h).strip() if h is not None else "" for h in header]
    out = []
    for row in rows:
        if not any(v not in (None, "") for v in row):
            continue
        out.append({n: v for n, v in zip(names, row) if n and v not in (None, "")})
    return out


def adapt_nextseek_workbook(path: Path) -> NormalizedInput:
    """A downloaded `*_AllMetadata*.xlsx`: one sheet per sample type. No API call."""
    path = Path(path)
    out = NormalizedInput(source={"adapter": "nextseek_workbook", "path": str(path)})
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in wb.sheetnames:
            for rec in _rows_from_sheet(wb[sheet]):
                uid = _clean(rec.get("UID")) or ""
                stype = (_clean(rec.get("SampleType"))
                         or _sample_type_from_uid(uid)
                         or sheet)
                out.samples.append(NormalizedSample(
                    sample_type=stype, uid=uid, metadata=rec,
                    parent=_clean(rec.get("Parent")),
                ))
    finally:
        wb.close()
    return out


def adapt_curated_sheet(path: Path) -> NormalizedInput:
    """A curated flat `Arm{X}.xlsx`. Works BEFORE upload - no API, no network.

    That matters because GEO deposit happens before NExtSEEK upload: accessions
    must be backfilled into the sheets first.
    """
    path = Path(path)
    out = NormalizedInput(source={"adapter": "curated_sheet", "path": str(path)})
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = "Samples" if "Samples" in wb.sheetnames else wb.sheetnames[0]
        for rec in _rows_from_sheet(wb[sheet]):
            uid = _clean(rec.get("uid")) or _clean(rec.get("UID")) or ""
            raw = rec.get("json_metadata")
            meta: dict = {}
            if raw:
                try:
                    meta = json.loads(raw)
                except (TypeError, ValueError) as exc:
                    meta = {"_json_metadata_error": f"{type(exc).__name__}: {exc}"}
                if not isinstance(meta, dict):
                    meta = {"_json_metadata_error": f"non-object JSON: {type(meta).__name__}"}
            # Denormalized columns fill anything json_metadata lacks.
            for col, key in (("name", "Name"), ("parent", "Parent")):
                if rec.get(col) and not meta.get(key):
                    meta[key] = rec[col]
            meta.setdefault("UID", uid)
            out.samples.append(NormalizedSample(
                sample_type=(_clean(rec.get("sampletype"))
                             or _clean(rec.get("sample_type"))
                             or _sample_type_from_uid(uid)),
                uid=uid,
                metadata=meta,
                parent=_clean(rec.get("parent")) or _clean(meta.get("Parent")),
            ))
    finally:
        wb.close()
    return out


def adapt_tabular(path: Path, *, sample_type: str | None = None) -> NormalizedInput:
    """Arbitrary xlsx or csv. Columns get mapped by the LLM step downstream."""
    path = Path(path)
    out = NormalizedInput(source={"adapter": "tabular", "path": str(path)})
    if path.suffix.lower() == ".csv":
        with path.open(newline="") as f:
            records = [dict(r) for r in csv.DictReader(f)]
    else:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            records = _rows_from_sheet(wb[wb.sheetnames[0]])
        finally:
            wb.close()
    for rec in records:
        uid = _clean(rec.get("UID")) or _clean(rec.get("uid")) or ""
        out.samples.append(NormalizedSample(
            sample_type=(sample_type or _clean(rec.get("SampleType"))
                         or _sample_type_from_uid(uid)),
            uid=uid, metadata=rec, parent=_clean(rec.get("Parent")),
        ))
    return out


# ── dispatch ───────────────────────────────────────────────────────────────

def detect_adapter(target) -> str:
    """Pick an adapter from what the curator handed us."""
    if isinstance(target, (list, tuple)):
        return "uids"
    p = Path(target)
    name = p.name
    if name.upper() == "RETRIEVE.TXT":
        return "retrieve_txt"
    if "AllMetadata" in name:
        return "nextseek_workbook"
    if p.suffix.lower() == ".xlsx" and name.startswith("Arm") and "_" not in p.stem:
        return "curated_sheet"
    return "tabular"


def adapt(target, **kwargs) -> NormalizedInput:
    which = detect_adapter(target)
    return {
        "uids": lambda: adapt_uids(list(target), fetch=kwargs.get("fetch")),
        "retrieve_txt": lambda: adapt_retrieve_txt(target, fetch=kwargs.get("fetch")),
        "nextseek_workbook": lambda: adapt_nextseek_workbook(target),
        "curated_sheet": lambda: adapt_curated_sheet(target),
        "tabular": lambda: adapt_tabular(target,
                                         sample_type=kwargs.get("sample_type")),
    }[which]()


# ── lineage ────────────────────────────────────────────────────────────────

def index_by_uid(normalized: NormalizedInput) -> dict[str, NormalizedSample]:
    return {s.uid: s for s in normalized.samples if s.uid}


def resolve_via_lineage(sample: NormalizedSample,
                        by_uid: dict[str, NormalizedSample],
                        key: str, *, max_depth: int = _MAX_LINEAGE_DEPTH) -> str | None:
    """Find `key` on this sample, else walk the Parent chain upward.

    Leaf wins: a value already on the sample is never overwritten by an
    ancestor's. Organism, tissue and cell line frequently live on ancestors
    rather than the D.SEQ row, which is what makes this necessary.

    Cycle-safe and depth-bounded: legacy PI data has many-to-many parents and
    is not guaranteed acyclic.
    """
    seen: set[str] = set()
    frontier = [sample]
    depth = 0
    while frontier and depth <= max_depth:
        nxt: list[NormalizedSample] = []
        for node in frontier:
            if node.uid in seen:
                continue
            seen.add(node.uid)
            value = node.metadata.get(key)
            if value not in (None, ""):
                return str(value)
            for parent_uid in str(node.parent or "").split(";"):
                parent_uid = parent_uid.strip()
                if parent_uid and parent_uid in by_uid and parent_uid not in seen:
                    nxt.append(by_uid[parent_uid])
        frontier = nxt
        depth += 1
    return None
