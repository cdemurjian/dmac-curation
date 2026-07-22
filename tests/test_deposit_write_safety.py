"""Deposit scripts must default to dry-run and mutate only under --write."""
import subprocess
import sys
from pathlib import Path

import pytest

REPO = Path(__file__).resolve().parent.parent

# Make the scripts/ dir importable so the in-place patchers can be driven
# behaviorally (their apply() entrypoints), not just scanned as source.
sys.path.insert(0, str(REPO / "scripts"))

DEPOSIT_SCRIPTS = [
    "scripts/stage_zenodo.py",
    "scripts/apply_zenodo_links.py",
    "scripts/apply_geo_accessions.py",
    "scripts/apply_omero_ids.py",
    # Added post-Task-3: smb_pull.py also had --dry-run, gating a network pull.
    "scripts/smb_pull.py",
]

# Import the hardened detector rather than re-declaring a regex here.
# The naive `add_argument\(\s*["\'](--[a-z0-9-]+)` form FAILS OPEN on a
# short-form-first declaration -- `add_argument("-n", "--dry-run")` parses as
# [] -- which would make this file, where write-safety is the entire subject,
# ship with a guard that silently passes. Task 3's reviewer proved that against
# 13 declaration forms. pytest puts the tests dir on sys.path, so this import
# works without packaging.
from test_curate_commands_present import parsed_flags  # noqa: E402


def _flags(rel: str) -> set[str]:
    return parsed_flags(REPO / rel)


@pytest.mark.parametrize("rel", DEPOSIT_SCRIPTS)
def test_has_write_flag(rel):
    assert "--write" in _flags(rel), f"{rel} must expose --write"


@pytest.mark.parametrize("rel", DEPOSIT_SCRIPTS)
def test_has_no_dry_run_flag(rel):
    assert "--dry-run" not in _flags(rel), (
        f"{rel} still has --dry-run; its absence implies writing"
    )


@pytest.mark.parametrize("rel", DEPOSIT_SCRIPTS)
def test_help_documents_default_is_dry_run(rel):
    result = subprocess.run(
        ["uv", "run", "--script", str(REPO / rel), "--help"],
        capture_output=True, text=True, timeout=120,
    )
    assert result.returncode == 0, result.stderr
    assert "default is dry-run" in result.stdout, (
        f"{rel} --help must state 'default is dry-run' on the --write flag"
    )


def test_command_doc_states_the_write_convention():
    doc = (REPO / "commands" / "curate-deposit.md").read_text()
    assert "--dry-run" not in doc, (
        "curate-deposit.md still documents --dry-run"
    )
    assert "default to dry-run and require `--write`" in doc


def test_zenodo_backfill_flag_matches_its_command_doc():
    """curate-deposit.md:22 documents --record-id; the script had
    --zenodo-record. Renamed so the doc and the CLI agree."""
    flags = _flags("scripts/apply_zenodo_links.py")
    assert "--record-id" in flags
    assert "--zenodo-record" not in flags, (
        "old name still present; a two-name CLI is how drift restarts"
    )


def test_smb_pull_converted_too():
    """smb_pull.py's --dry-run gated an actual network transfer."""
    flags = _flags("scripts/smb_pull.py")
    assert "--write" in flags
    assert "--dry-run" not in flags


def test_no_script_anywhere_still_uses_dry_run():
    """Repo-wide sweep, so a fourth offender cannot hide."""
    offenders = [
        p.relative_to(REPO) for p in (REPO / "scripts").rglob("*.py")
        if "--dry-run" in _flags(str(p.relative_to(REPO)))
    ]
    assert not offenders, f"still using --dry-run: {offenders}"


# ----------------------------------------------------------------------------
# Backup-before-in-place-overwrite regression guards (data-loss fix).
#
# apply_omero_ids.py and apply_zenodo_links.py overwrite the curator's real
# upload sheets in place. Their sibling apply_geo_accessions.py has always
# written a .bak first; these two did not, so a wrong --record-id / mismatched
# OMERO CSV run with --write would clobber Link_PrimaryData with no recovery.
# The omero path is driven behaviorally (its apply() takes plain paths); the
# zenodo path (which needs project-root config + a zip dir) is guarded at the
# source level so the copy-before-save cannot silently regress.
# ----------------------------------------------------------------------------

def _make_omero_sheet(path: Path) -> None:
    """A minimal upload sheet: File_PrimaryData populated, Link_PrimaryData empty."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Samples"
    ws.append(["File_PrimaryData", "Link_PrimaryData"])
    ws.append(["img1.tif", None])
    wb.save(path)
    wb.close()


def _link_value(path: Path):
    """Read back the single Samples!Link_PrimaryData cell (row 2).

    Opened via a binary handle so a `.bak`-suffixed copy (still a real xlsx
    payload) loads despite openpyxl's by-path extension check.
    """
    from openpyxl import load_workbook

    with path.open("rb") as fh:
        wb = load_workbook(fh, data_only=True)
        try:
            return wb["Samples"].cell(row=2, column=2).value
        finally:
            wb.close()


def test_apply_omero_backup_created_on_write(tmp_path):
    """--write must drop a .bak holding the pre-write content before overwriting."""
    import apply_omero_ids

    xlsx = tmp_path / "A.IMG-upload-new.xlsx"
    _make_omero_sheet(xlsx)

    csv_path = tmp_path / "omero_images.csv"
    csv_path.write_text("filename,web_url\nimg1.tif,https://omero.example/img1\n")

    bak = xlsx.with_suffix(xlsx.suffix + ".bak")
    assert not bak.exists()

    rc = apply_omero_ids.apply(xlsx, csv_path, dry_run=False)
    assert rc == 0

    # The live sheet was patched...
    assert _link_value(xlsx) == "https://omero.example/img1"
    # ...and a backup exists preserving the pre-write (empty) Link_PrimaryData.
    assert bak.exists(), "no .bak created before in-place overwrite"
    assert _link_value(bak) is None, ".bak does not hold the pre-write content"
    # The backup genuinely differs from the overwritten live sheet.
    assert _link_value(bak) != _link_value(xlsx)


def test_apply_omero_no_backup_on_dry_run(tmp_path):
    """Dry-run must NOT touch disk: no overwrite, no .bak."""
    import apply_omero_ids

    xlsx = tmp_path / "A.IMG-upload-new.xlsx"
    _make_omero_sheet(xlsx)

    csv_path = tmp_path / "omero_images.csv"
    csv_path.write_text("filename,web_url\nimg1.tif,https://omero.example/img1\n")

    rc = apply_omero_ids.apply(xlsx, csv_path, dry_run=True)
    assert rc == 0

    bak = xlsx.with_suffix(xlsx.suffix + ".bak")
    assert not bak.exists(), "dry-run must not create a .bak"
    assert _link_value(xlsx) is None, "dry-run must not modify the live sheet"


def test_apply_zenodo_backs_up_before_in_place_save():
    """Source guard: the in-place wb.save must be preceded by a shutil.copy .bak.

    apply_zenodo_links needs project-root config + a zip dir to drive, so this
    guards the write path structurally: the only `wb.save(sheet_path)` must be
    immediately preceded by a `shutil.copy(... ".bak" ...)` of the same path.
    """
    src = (REPO / "scripts" / "apply_zenodo_links.py").read_text()
    assert "import shutil" in src, "apply_zenodo_links.py must import shutil"

    lines = src.splitlines()
    save_idxs = [i for i, ln in enumerate(lines) if "wb.save(sheet_path)" in ln]
    assert save_idxs, "expected an in-place wb.save(sheet_path)"
    for i in save_idxs:
        prev = lines[i - 1]
        assert "shutil.copy(sheet_path" in prev and '".bak"' in prev, (
            "the line before `wb.save(sheet_path)` must copy a .bak backup; "
            f"got: {prev!r}"
        )
